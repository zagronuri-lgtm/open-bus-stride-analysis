"""Tests for branch 500 ground-truth workbook (Power BI first)."""

from __future__ import annotations

import json
from pathlib import Path

from src.branch_500_ground_truth import BI_TRUTH, build_workbook, load_mapping_500


def test_bi_truth_anchors_in_2_to_3_percent_band():
    pct = BI_TRUTH["aggregate_branch_500"]["pct"]
    assert 0.02 <= pct <= 0.03
    assert BI_TRUTH["day_2026_07_05"]["pct"] == 0.0531


def test_mapping_500_count():
    rows = load_mapping_500()
    assert len(rows) == 20
    assert {r["route_mkt"] for r in rows} >= {"11501", "11230", "11231"}
    # קווי 50x לא ממופים לסניף אחר
    assert all(
        (r.get("route_short_name") or "").startswith("50") is False
        or r["branch"] == "500"
        for r in rows
    )
    # אין שיוך לסניף חולון/נתניה בתוך רשימת 500
    assert all(r["branch"] == "500" for r in rows)
    clusters = {r["cluster"] for r in rows}
    assert clusters <= {"שרון", "שרון חולון מרחבי"}


def test_ground_truth_workbook_builds(tmp_path: Path):
    inv = {
        "summary": {
            "windows": {
                "valid_weekdays_sun_tue": {
                    "planned": 3866,
                    "executed": 3850,
                    "missing": 16,
                    "pct": 0.00414,
                },
                "weekdays_sun_thu_incl_segment": {
                    "planned": 6472,
                    "executed": 6412,
                    "missing": 60,
                    "pct": 0.00927,
                },
                "full_week": {
                    "planned": 7223,
                    "executed": 7146,
                    "missing": 77,
                    "pct": 0.01066,
                },
                "july_only_wed_thu": {
                    "planned": 2606,
                    "executed": 2562,
                    "missing": 44,
                    "pct": 0.01688,
                },
            },
            "by_day": {
                "2026-06-28": {
                    "planned": 1286,
                    "executed": 1285,
                    "missing": 1,
                    "pct": 0.0008,
                    "valid": True,
                    "classification": "תקין",
                }
            },
            "by_mkt_valid": {
                "11230": {
                    "planned": 10,
                    "executed": 2,
                    "missing": 8,
                    "line_ref": "7728",
                    "pct": 0.8,
                }
            },
        }
    }
    out = tmp_path / "truth.xlsx"
    path = build_workbook(inv, out)
    assert path.exists()
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert wb.sheetnames[0] == "נתוני אמת — BI"
    assert "הדבק BI חי" in wb.sheetnames
    assert "השוואת Stride (משני)" in wb.sheetnames
    assert "מיפוי 20 מק״ט" in wb.sheetnames
    assert "דוגמאות פער" in wb.sheetnames
    assert "מסקנות חקירה" in wb.sheetnames
