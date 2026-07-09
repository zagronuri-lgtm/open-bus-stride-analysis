"""
חוברת כיול: Open Bus Stride (אי-יציאה 2.1.1) מול Power BI (אמת רשמית).

הרצה:
  python -m src.powerbi_calibration \\
    --weekly-xlsx outputs/דוח_שבועי_אי_ביצוע_2026-07-04.xlsx \\
    --output-dir outputs
"""

from __future__ import annotations

import argparse
import datetime as dt
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

POWERBI_URL = (
    "https://app.powerbi.com/groups/me/reports/"
    "cae01515-0e87-42dc-b643-244528bc2bca/bf6381d728e43ccdcd08"
    "?experience=power-bi"
)

# עוגנים מכיול 2026-07-08 (docs/powerbi_ground_truth.md)
BI_ANCHORS = {
    "2026-04": {
        "label": "אפריל 2026 (חודש)",
        "latest_safe_date": "2026-05-01",
        "rides": 237_977,
        "rides_for_calc": 189_634,
        "exclusions": 41_885,
        "non_execution": 1_528,
        "delay": 3_376,
        "early": 101,
        "branch": "All",
        "cluster": "All",
    },
    "2026-07_partial_to_06": {
        "label": "יולי 2026 גלוי עד 06/07 (כיול 08/07)",
        "latest_safe_date": "2026-07-06",
        "rides": 57_145,
        "rides_for_calc": 56_392,
        "exclusions": None,
        "non_execution": 1_108,
        "delay": 1_213,
        "early": 44,
        "branch": "All",
        "cluster": "All",
    },
    # נשלף חי מ-Power BI ב-2026-07-09 דרך browser MCP
    "2026-07_partial": {
        "label": "יולי 2026 גלוי עד 07/07",
        "latest_safe_date": "2026-07-07",
        "rides": 69_644,
        "rides_for_calc": 68_717,
        "exclusions": None,
        "non_execution": 1_234,
        "delay": 1_472,
        "early": 54,
        "last_trip": 6,
        "low_freq": 11,
        "unauthorized": 9,
        "branch": "All",
        "cluster": "All",
        "source": "live Power BI 2026-07-09; Month=2026-07 Branch=All Cluster slicer=All",
    },
}

# עוגני סניף 500 (כיול 08/07 + חקירת 09/07) — אין שליפה חיה Branch=500 בסביבת cloud
BI_BRANCH_500_ANCHORS = {
    "aggregate": {
        "label": "אגרגט סניף 500 (כיול 08/07)",
        "pct": 0.0253,
        "note": "METROPOLIN_BRANCHES_AGG; משתמש מדווח 2–3% קבוע",
    },
    "day_2026-07-05": {
        "label": "יום 05/07",
        "pct": 0.0531,
        "note": "Stride=BI=5.31% באותו יום",
    },
}

# מיפוי שמות סניף: מקומי (קובץ מק״ט) ↔ BI (כיול קודם)
BRANCH_NAME_MAP = [
    # (local_name, bi_name, note)
    ("500", "500", "תואם"),
    ("אלעד", "אלעד", "תואם"),
    ("אונו", "אונו", "תואם"),
    ("שרון עירוני", "שרון עירוני", "תואם"),
    ("שרון בינעירוני", "שרון ב.ע", "לוודא מול BI"),
    ('ב"ש', "הנגב", "לוודא — ייתכן שם אשכול ב-BI"),
    ("חולון", "חולון / שח\"מ?", "לבדוק ב-BI"),
    ("נתניה", "נתניה / שח\"מ?", "לבדוק ב-BI"),
    ("אלעד - לא ברישוי", "—", "מקומי בלבד"),
    ("לא משויך", "—", "פער מיפוי מק״ט"),
]


def _styles():
    return {
        "arial": "Arial",
        "hdr_fill": PatternFill("solid", fgColor="1F4E78"),
        "hdr_font": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "title": Font(name="Arial", bold=True, size=14, color="1F4E78"),
        "base": Font(name="Arial", size=10),
        "warn": PatternFill("solid", fgColor="FCE4D6"),
        "ok": PatternFill("solid", fgColor="E2EFDA"),
        "bad": PatternFill("solid", fgColor="F8CBAD"),
        "thin": Border(
            left=Side(style="thin", color="BFBFBF"),
            right=Side(style="thin", color="BFBFBF"),
            top=Side(style="thin", color="BFBFBF"),
            bottom=Side(style="thin", color="BFBFBF"),
        ),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _header_row(ws, headers, row=1, styles=None):
    s = styles or _styles()
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = s["hdr_font"]
        c.fill = s["hdr_fill"]
        c.alignment = s["center"]
        c.border = s["thin"]
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, widths):
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w


def read_stride_metropoline(weekly_xlsx: Path) -> dict:
    """קורא סיכום מטרופולין + לפי סניף מהדוח השבועי."""
    wb = load_workbook(weekly_xlsx, data_only=False)
    out: dict = {
        "path": str(weekly_xlsx),
        "operator": {},
        "branches": [],
        "valid_days_note": "",
    }

    if "סיכום מנהלים" in wb.sheetnames:
        ws = wb["סיכום מנהלים"]
        for row in ws.iter_rows(min_row=5, max_row=12, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0])
            if name == "מטרופולין":
                planned, executed = int(row[1] or 0), int(row[2] or 0)
                missing = planned - executed
                out["operator"] = {
                    "name": name,
                    "planned": planned,
                    "executed": executed,
                    "missing": missing,
                    "missing_pct": (missing / planned) if planned else 0.0,
                    "valid_days": row[7],
                }
            if name == "סך הכל":
                break

    if "מטרופולין לפי סניף" in wb.sheetnames:
        ws = wb["מטרופולין לפי סניף"]
        for row in ws.iter_rows(min_row=5, max_row=30, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0])
            if name.startswith("סך"):
                continue
            planned = int(row[1] or 0)
            executed = int(row[2] or 0)
            missing = int(row[3] or 0) if len(row) > 3 and row[3] is not None else planned - executed
            out["branches"].append({
                "branch": name,
                "planned": planned,
                "executed": executed,
                "missing": missing,
                "missing_pct": (missing / planned) if planned else 0.0,
                "mkts": row[7] if len(row) > 7 else None,
            })

    if "ימים חריגים" in wb.sheetnames:
        ws = wb["ימים חריגים"]
        notes = []
        for row in ws.iter_rows(min_row=4, max_row=12, values_only=True):
            if not row or not row[0]:
                continue
            notes.append(f"{row[0]} {row[1]}: {row[2]} ({row[3]})")
        out["valid_days_note"] = " | ".join(notes)

    # week label from filename
    stem = weekly_xlsx.stem
    if "2026-" in stem:
        out["week_ending"] = stem.split("_")[-1]
    return out


def build_calibration_workbook(stride: dict, output_path: Path) -> Path:
    s = _styles()
    wb = Workbook()
    wb.remove(wb.active)

    # ---- 1. הוראות ----
    ws = wb.create_sheet("הוראות כיול")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "כיול אי-ביצוע: Stride (2.1.1) מול Power BI"
    ws["A1"].font = s["title"]
    lines = [
        ("קישור Power BI", POWERBI_URL),
        ("עמוד", "נסיעות"),
        ("לפני קריאה", "Reset to default → Month / Branch / Cluster"),
        ("% BI רשמי", "אי ביצוע ÷ נסיעות לחישוב"),
        ("% Stride", "אי-בוצע ÷ מתוכנן (אי-יציאה בלבד = חסם תחתון)"),
        ("פער צפוי", "BI ≥ Stride אם BI כולל איחור/הקדמה חמורים והחרגות שונות"),
        ("מקור Stride", stride.get("path", "")),
        ("שבוע", stride.get("week_ending", "")),
        ("ימים בדוח", stride.get("valid_days_note", "")),
        ("מסמך", "docs/powerbi_ground_truth.md"),
    ]
    _header_row(ws, ["פריט", "ערך"], row=3, styles=s)
    for i, (k, v) in enumerate(lines, start=4):
        ws.cell(i, 1, k).font = Font(name="Arial", bold=True, color="1F4E78")
        ws.cell(i, 2, v).font = s["base"]
        ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="center")
    _autosize(ws, [22, 100])

    # ---- 2. הדבק BI ----
    ws = wb.create_sheet("הדבק KPI מ-BI")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "הדבק כאן ערכים מהעמוד נסיעות ב-Power BI (לאחר סינון)"
    ws["A1"].font = s["title"]
    _header_row(ws, ["שדה", "ערך", "הערה"], row=3, styles=s)
    paste_fields = [
        ("תקופה (Month)", "", "למשל 2026-07"),
        ("Branch", "All", ""),
        ("Cluster", "All", ""),
        ("תאריך כספת אחרון", "", "latest safe date"),
        ("נסיעות", "", "אריח KPI — שורה זו = B8"),
        ("נסיעות לחישוב", "", "מכנה — שורה זו = B9"),
        ("החרגות", "", "אם מוצג — שורה זו = B10"),
        ("אי ביצוע", "", "מונה — שורה זו = B11"),
        ("איחור", "", "אופציונלי"),
        ("הקדמה", "", "אופציונלי"),
        ("% רשמי BI", None, "אי ביצוע / נסיעות לחישוב"),
        ("מקדם מכנה", None, "נסיעות לחישוב / נסיעות"),
    ]
    for i, (field, default, note) in enumerate(paste_fields, start=4):
        ws.cell(i, 1, field).font = Font(name="Arial", bold=True)
        ws.cell(i, 2, default)
        ws.cell(i, 3, note).font = Font(name="Arial", italic=True, size=9, color="808080")
        for j in range(1, 4):
            ws.cell(i, j).border = s["thin"]
    # row 14 = % רשמי, row 15 = מקדם מכנה (4+10, 4+11)
    ws["B14"] = '=IF(OR(B9="",B9=0),"",B11/B9)'
    ws["B14"].number_format = "0.00%"
    ws["B15"] = '=IF(OR(B8="",B8=0),"",B9/B8)'
    ws["B15"].number_format = "0.000"
    _autosize(ws, [22, 18, 40])
    ws["A17"] = "אחרי מילוי — השווה לגיליון השוואת מטרופולין"
    ws["A17"].font = Font(name="Arial", italic=True, color="C65911")

    # ---- 3. עוגני BI שמורים ----
    ws = wb.create_sheet("עוגני BI שמורים")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "עוגנים מכיול קודם (2026-07-08) — לא לרענון אוטומטי"
    ws["A1"].font = s["title"]
    _header_row(
        ws,
        ["מפתח", "תווית", "תאריך כספת", "נסיעות", "נסיעות לחישוב", "החרגות",
         "אי ביצוע", "% רשמי", "מקדם מכנה", "Branch", "Cluster"],
        row=3, styles=s,
    )
    rr = 4
    for key, a in BI_ANCHORS.items():
        rides = a["rides"]
        rfc = a["rides_for_calc"]
        ne = a["non_execution"]
        ws.cell(rr, 1, key)
        ws.cell(rr, 2, a["label"])
        ws.cell(rr, 3, a["latest_safe_date"])
        ws.cell(rr, 4, rides).number_format = "#,##0"
        ws.cell(rr, 5, rfc).number_format = "#,##0"
        ws.cell(rr, 6, a.get("exclusions") if a.get("exclusions") is not None else "—")
        ws.cell(rr, 7, ne).number_format = "#,##0"
        ws.cell(rr, 8, ne / rfc if rfc else 0).number_format = "0.00%"
        ws.cell(rr, 9, rfc / rides if rides else 0).number_format = "0.000"
        ws.cell(rr, 10, a.get("branch", ""))
        ws.cell(rr, 11, a.get("cluster", ""))
        rr += 1
    _autosize(ws, [18, 28, 14, 12, 14, 12, 12, 10, 12, 10, 10])

    # ---- 4. השוואת מטרופולין ----
    ws = wb.create_sheet("השוואת מטרופולין")
    ws.sheet_view.rightToLeft = True
    op = stride.get("operator") or {}
    ws["A1"] = "מטרופולין — Stride מול BI (מלא ידנית את עמודת BI)"
    ws["A1"].font = s["title"]
    ws["A2"] = (
        f"Stride שבוע {stride.get('week_ending', '')}: "
        f"מתוכנן={op.get('planned', 0):,} | בוצע={op.get('executed', 0):,} | "
        f"אי-בוצע={op.get('missing', 0):,} | %={op.get('missing_pct', 0):.2%} | "
        f"ימי חול תקפים={op.get('valid_days', '')}"
    )
    ws["A2"].font = Font(name="Arial", size=9, color="666666")

    _header_row(
        ws,
        ["מדד", "Stride (2.1.1)", "Power BI", "פער (BI−Stride)", "הערה"],
        row=4, styles=s,
    )
    rows_cmp = [
        ("נסיעות / מתוכנן", op.get("planned"), None, "מכנה Stride = מתוכנן; BI = נסיעות או נסיעות לחישוב"),
        ("בוצע / נסיעות לחישוב", op.get("executed"), None, "לא אותה הגדרה — להשוות בזהירות"),
        ("אי-ביצוע (מונה)", op.get("missing"), None, "BI כולל רכיבים מעבר לאי-יציאה"),
        ("% אי-ביצוע", op.get("missing_pct"), None, "% BI = אי-ביצוע/נסיעות לחישוב"),
    ]
    for i, (metric, stride_val, _bi, note) in enumerate(rows_cmp, start=5):
        ws.cell(i, 1, metric)
        cell = ws.cell(i, 2, stride_val)
        if metric.startswith("%"):
            cell.number_format = "0.00%"
        else:
            cell.number_format = "#,##0"
        ws.cell(i, 3, None)  # paste BI
        ws.cell(i, 3).fill = s["warn"]
        if metric.startswith("%"):
            ws.cell(i, 4, f'=IF(OR(C{i}="",B{i}=""),"",C{i}-B{i})').number_format = "0.00%"
        else:
            ws.cell(i, 4, f'=IF(OR(C{i}="",B{i}=""),"",C{i}-B{i})').number_format = "#,##0"
        ws.cell(i, 5, note).font = Font(name="Arial", italic=True, size=9, color="808080")
    ws["A10"] = "תאים כתומים = להדביק מ-Power BI לאותה תקופה (או לציין אי-התאמת חלון זמן)"
    ws["A10"].font = Font(name="Arial", italic=True, color="C65911")
    _autosize(ws, [26, 16, 14, 16, 55])

    # ---- 5. לפי סניף ----
    ws = wb.create_sheet("סניפים Stride מול BI")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "מטרופולין לפי סניף — Stride מהדוח השבועי; BI להדבקה"
    ws["A1"].font = s["title"]
    _header_row(
        ws,
        ["סניף מקומי", "שם ב-BI (מיפוי)", "מתוכנן Stride", "אי-בוצע Stride",
         "% Stride", "% BI (הדבק)", "פער נק׳ %", "מק״טים", "הערת מיפוי"],
        row=3, styles=s,
    )
    bi_lookup = {local: (bi, note) for local, bi, note in BRANCH_NAME_MAP}
    rr = 4
    for b in sorted(stride.get("branches") or [], key=lambda x: -x["planned"]):
        local = b["branch"]
        bi_name, note = bi_lookup.get(local, ("?", "לא במפת שמות — לבדוק ב-BI"))
        ws.cell(rr, 1, local)
        ws.cell(rr, 2, bi_name)
        ws.cell(rr, 3, b["planned"]).number_format = "#,##0"
        ws.cell(rr, 4, b["missing"]).number_format = "#,##0"
        ws.cell(rr, 5, b["missing_pct"]).number_format = "0.00%"
        ws.cell(rr, 6, None).fill = s["warn"]
        ws.cell(rr, 6).number_format = "0.00%"
        ws.cell(rr, 7, f'=IF(OR(F{rr}="",E{rr}=""),"",F{rr}-E{rr})').number_format = "0.00%"
        ws.cell(rr, 8, b.get("mkts"))
        ws.cell(rr, 9, note).font = Font(name="Arial", italic=True, size=9, color="808080")
        if local == "לא משויך" and b["planned"] > 0:
            for j in range(1, 10):
                if ws.cell(rr, j).fill != s["warn"]:
                    ws.cell(rr, j).fill = s["bad"]
        rr += 1
    _autosize(ws, [18, 18, 14, 14, 10, 12, 10, 10, 35])

    # ---- 6. מיפוי שמות ----
    ws = wb.create_sheet("מיפוי שמות סניף")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "טבלת תרגום שמות — לעדכן אחרי בדיקה ב-Power BI"
    ws["A1"].font = s["title"]
    _header_row(ws, ["שם מקומי (מק״ט→סניף)", "שם ב-Power BI", "סטטוס"], row=3, styles=s)
    for i, (local, bi, note) in enumerate(BRANCH_NAME_MAP, start=4):
        ws.cell(i, 1, local)
        ws.cell(i, 2, bi)
        ws.cell(i, 3, note)
    _autosize(ws, [28, 22, 40])

    # ---- 7. מתודולוגיה ----
    ws = wb.create_sheet("מתודולוגיה")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "מגבלות כיול"
    ws["A1"].font = s["title"]
    notes = [
        "Stride מודד אי-יציאה (2.1.1) בלבד — חסם תחתון לנספח כ\"ו.",
        "Power BI כולל לרוב רכיבי אי-ביצוע נוספים + החרגות תפעוליות (נסיעות לחישוב).",
        "חלון זמן חייב להיות זהה: שבוע מול שבוע, או חודש מול חודש — לא לערבב.",
        "בדוח השבועי הנוכחי ייתכנו ימים ב-segment (חופש גדול) שלא נכנסים לממוצע.",
        "בדף BI שנבדק אין סנן יום בודד — רק Month מצטבר עד תאריך כספת.",
        "שמות סניף ב-BI עשויים להיות מקוצרים (ב.ע / הנגב) — ראה גיליון מיפוי שמות.",
        "סניף 500: Stride א׳–ג׳ תקפים ≈0.41% מול עוגן BI ≈2.53%. הפער = חלון (חופש גדול) + הגדרה 2.1.1; המיפוי (20 מק״ט, 18 פעילים בימים תקפים) יציב. ביום 05/07 היה 5.31%=BI. אין באג ספירה — תועד + גיליון רגישות.",
        "פירוט חקירה: outputs/חקירת_סניף_500_*.xlsx ו-python -m src.investigate_branch_500",
    ]
    for i, t in enumerate(notes, start=3):
        ws.cell(i, 1, f"{i-2}.")
        ws.cell(i, 2, t).font = s["base"]
        ws.cell(i, 2).alignment = Alignment(wrap_text=True)
    _autosize(ws, [4, 100])

    # ---- 8. סניף 500 ----
    ws = wb.create_sheet("סניף 500 — פער")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "סניף 500 — Stride מול עוגני BI"
    ws["A1"].font = s["title"]
    _header_row(ws, ["מקור", "%", "הערה"], row=3, styles=s)
    branch500 = next((b for b in (stride.get("branches") or []) if b["branch"] == "500"), None)
    rr = 4
    if branch500:
        ws.cell(rr, 1, "Stride דוח שבועי (ימי חול תקפים)")
        ws.cell(rr, 2, branch500["missing_pct"]).number_format = "0.00%"
        ws.cell(
            rr, 3,
            f"מתוכנן={branch500['planned']:,} אי-בוצע={branch500['missing']}",
        )
        rr += 1
    for key, a in BI_BRANCH_500_ANCHORS.items():
        ws.cell(rr, 1, a["label"])
        ws.cell(rr, 2, a["pct"]).number_format = "0.00%"
        ws.cell(rr, 2).fill = s["warn"]
        ws.cell(rr, 3, a["note"])
        rr += 1
    rr += 1
    ws.cell(rr, 1, "מסקנה").font = Font(name="Arial", bold=True, color="1F4E78")
    ws.cell(
        rr, 2,
        "לא באג מיפוי. להשוות BI Branch=500 לאותו חלון יומי; ראה גיליון רגישות בדוח השבועי.",
    )
    _autosize(ws, [40, 12, 70])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    ap = argparse.ArgumentParser(description="חוברת כיול Stride מול Power BI")
    ap.add_argument("--weekly-xlsx", required=True, help="נתיב לדוח השבועי xlsx")
    ap.add_argument("--output-dir", default="outputs")
    args = ap.parse_args()

    weekly = Path(args.weekly_xlsx).expanduser()
    if not weekly.exists():
        raise SystemExit(f"לא נמצא: {weekly}")

    stride = read_stride_metropoline(weekly)
    week = stride.get("week_ending") or dt.date.today().isoformat()
    out = Path(args.output_dir) / f"כיול_Stride_מול_PowerBI_{week}.xlsx"
    path = build_calibration_workbook(stride, out)
    print(f"[✓] נשמר: {path}")
    op = stride.get("operator") or {}
    print(
        f"[i] מטרופולין Stride: מתוכנן={op.get('planned')} אי-בוצע={op.get('missing')} "
        f"%={op.get('missing_pct', 0):.2%} | סניפים={len(stride.get('branches') or [])}"
    )
    print(f"[i] Power BI: {POWERBI_URL}")


if __name__ == "__main__":
    main()
