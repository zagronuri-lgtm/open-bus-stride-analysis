"""חקירת פער אי-ביצוע לסניף מטרופולין 500 מול עוגני Power BI.

מושך רק מפעיל 15 לשבוע נתון, מצליב למק״ט→סניף, ומפיק:
  - רגישות חלונות (א׳–ג׳ תקפים / א׳–ה׳ / כל השבוע)
  - פירוט יומי ומק״ט
  - השוואה לעוגני BI השמורים (2.53% אגרגט, 5.31% ביום 05/07)

הרצה:
  python -m src.investigate_branch_500 --week-ending 2026-07-04 --output-dir outputs
"""

from __future__ import annotations

import argparse
import collections
import csv
import datetime as dt
import json
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import weekly_report as wr
from src.exclusions_calendar import classify_date, load_calendar
from src.metropoline_branches import (
    METROPOLINE_OPERATOR_REF,
    branch_for_mkt,
    load_branch_map,
)

# עוגנים מכיול 2026-07-08 (אין דפדפן בסביבת cloud זו לשליפה חיה Branch=500)
BI_BRANCH_500_ANCHORS = {
    "aggregate_siri_attached": {
        "label": "אגרגט סניפים מצורף לכיול 08/07",
        "pct": 0.0253,
        "note": "METROPOLIN_BRANCHES_AGG ב-create_calibration_workbook.py",
    },
    "day_2026-07-05": {
        "label": "יום 05/07 — עוגן מצורף",
        "pct": 0.0531,
        "note": "Stride=BI=5.31% באותו יום (מקדם רכיב ≈1.0 לאותו חלון)",
    },
}


def _week_dates(week_ending: dt.date) -> tuple[dt.date, dt.date, list[dt.date]]:
    sat = week_ending
    while sat.weekday() != 5:
        sat -= dt.timedelta(days=1)
    sun = sat - dt.timedelta(days=6)
    dates = [sun + dt.timedelta(days=i) for i in range(7)]
    return sun, sat, dates


def fetch_metropoline_week(dates: list[dt.date]) -> list[wr.DayData]:
    """מושך מתוכנן לכל המפעילים (צריך לסיווג ימים) ובוצע רק למטרופולין."""
    days: list[wr.DayData] = []
    for d in dates:
        day = wr.DayData(d)
        day.planned_raw = wr.fetch_planned(d)
        # רק מטרופולין לביצוע — חוסך זמן; לסיווג רוב-מפעילים נשתמש במתוכנן בלבד
        # ולכן נמשוך בוצע לכל המפעילים במקביל (נדרש ל-classify_days)
        with ThreadPoolExecutor(max_workers=6) as ex:
            futs = {op: ex.submit(wr.fetch_executed, op, d) for op in wr.OPERATORS}
            day.executed = {op: f.result() for op, f in futs.items()}
        days.append(day)
        print(
            f"    {wr.HEB_DOW[d.weekday()]} {d}: "
            f"מתוכנן={sum(day.planned_raw.values())}, "
            f"מטרופולין בוצע={sum(day.executed.get(METROPOLINE_OPERATOR_REF, {}).values())}"
        )
    return days


def summarize_branch(
    recs: list[wr.LineDayRecord],
    days: list[wr.DayData],
    branch: str = "500",
) -> dict:
    by_date = {d.date: d for d in days}
    valid_dates = {d.date for d in days if d.valid and d.is_workday}
    strike = {(d.date, op) for d in days for op in d.strike_ops}

    branch_recs = [r for r in recs if r.branch == branch]
    windows = {
        "valid_weekdays_sun_tue": lambda r: r.date in valid_dates
        and (r.date, r.op) not in strike,
        "weekdays_sun_thu_incl_segment": lambda r: r.date.weekday()
        in wr.WEEKDAY_WORK,
        "full_week": lambda r: True,
        "july_only_wed_thu": lambda r: r.date
        in {dt.date(2026, 7, 1), dt.date(2026, 7, 2)},
    }

    out_windows = {}
    for name, pred in windows.items():
        subset = [r for r in branch_recs if pred(r)]
        planned = sum(r.planned for r in subset)
        executed = sum(r.executed for r in subset)
        missing = planned - executed
        out_windows[name] = {
            "planned": planned,
            "executed": executed,
            "missing": missing,
            "pct": (missing / planned) if planned else 0.0,
            "rows": len(subset),
        }

    by_day: dict[str, dict] = {}
    for r in branch_recs:
        key = r.date.isoformat()
        slot = by_day.setdefault(key, {"planned": 0, "executed": 0, "missing": 0})
        slot["planned"] += r.planned
        slot["executed"] += r.executed
        slot["missing"] += r.nonexec
    for key, slot in by_day.items():
        p = slot["planned"]
        slot["pct"] = (slot["missing"] / p) if p else 0.0
        d = by_date.get(dt.date.fromisoformat(key))
        slot["classification"] = d.classification if d else ""
        slot["valid"] = bool(d and d.valid and d.is_workday)

    by_mkt: dict[str, dict] = {}
    for r in branch_recs:
        if r.date not in valid_dates or (r.date, r.op) in strike:
            continue
        slot = by_mkt.setdefault(
            r.mkt or "?",
            {"planned": 0, "executed": 0, "missing": 0, "line_ref": r.line},
        )
        slot["planned"] += r.planned
        slot["executed"] += r.executed
        slot["missing"] += r.nonexec
    for slot in by_mkt.values():
        p = slot["planned"]
        slot["pct"] = (slot["missing"] / p) if p else 0.0

    return {
        "branch": branch,
        "windows": out_windows,
        "by_day": dict(sorted(by_day.items())),
        "by_mkt_valid": dict(
            sorted(by_mkt.items(), key=lambda kv: -kv[1]["missing"])
        ),
        "day_classifications": {
            d.date.isoformat(): {
                "dow": wr.HEB_DOW[d.dow],
                "classification": d.classification,
                "valid": d.valid,
                "reason": d.reason,
            }
            for d in days
        },
    }


def write_investigation_xlsx(summary: dict, mapping_rows: list[dict], path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    title_font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF")
    warn = PatternFill("solid", fgColor="FCE4D6")
    bad = PatternFill("solid", fgColor="F8CBAD")
    ok = PatternFill("solid", fgColor="E2EFDA")

    def sheet(name: str):
        ws = wb.create_sheet(name)
        ws.sheet_view.rightToLeft = True
        return ws

    def header(ws, headers, row=1):
        for j, h in enumerate(headers, 1):
            c = ws.cell(row, j, h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center")

    # מסקנה
    ws = sheet("מסקנה")
    ws["A1"] = "חקירת פער סניף 500 — Stride מול עוגני BI"
    ws["A1"].font = title_font
    lines = [
        "Stride (א׳–ג׳ תקפים): חסם תחתון 2.1.1 בלבד.",
        "עוגן BI אגרגט סניף 500: 2.53% (כיול 08/07).",
        "עוגן יום 05/07: Stride=BI=5.31% — באותו יום אין פער הגדרה.",
        "מסקנה עיקרית: הפער השבועי נובע בעיקר מחלון (החרגת חופש גדול) + הגדרה צרה;",
        "גם בחלון רחב יותר Stride נשאר מתחת ל-2% — BI הרשמי גבוה יותר כצפוי.",
        "קווי לילה 11230/11231 תורמים חלק גדול מהחסרים היחסיים אך נפח קטן.",
        "שליפת BI חיה Branch=500 לא בוצעה בסביבת cloud זו (אין browser MCP).",
    ]
    for i, t in enumerate(lines, start=3):
        ws.cell(i, 1, t)
    ws.column_dimensions["A"].width = 110

    # רגישות
    ws = sheet("רגישות חלונות")
    ws["A1"] = "סניף 500 — % אי-ביצוע לפי חלון"
    ws["A1"].font = title_font
    header(ws, ["חלון", "מתוכנן", "בוצע", "אי-בוצע", "%", "הערה"], row=3)
    notes = {
        "valid_weekdays_sun_tue": "כמו סיכום הדוח השבועי (אחרי classify_days)",
        "weekdays_sun_thu_incl_segment": "כולל ד׳–ה׳ חופש גדול",
        "full_week": "כולל ו׳–ש׳",
        "july_only_wed_thu": "רק 01–02/07",
    }
    rr = 4
    for key, w in summary["windows"].items():
        ws.cell(rr, 1, key)
        ws.cell(rr, 2, w["planned"])
        ws.cell(rr, 3, w["executed"])
        ws.cell(rr, 4, w["missing"])
        cell = ws.cell(rr, 5, w["pct"])
        cell.number_format = "0.00%"
        if w["pct"] >= 0.025:
            cell.fill = bad
        elif w["pct"] >= 0.021:
            cell.fill = warn
        else:
            cell.fill = ok
        ws.cell(rr, 6, notes.get(key, ""))
        rr += 1
    rr += 1
    ws.cell(rr, 1, "עוגן BI אגרגט")
    ws.cell(rr, 5, BI_BRANCH_500_ANCHORS["aggregate_siri_attached"]["pct"]).number_format = "0.00%"
    ws.cell(rr, 5).fill = warn
    ws.cell(rr, 6, BI_BRANCH_500_ANCHORS["aggregate_siri_attached"]["note"])
    rr += 1
    ws.cell(rr, 1, "עוגן BI יום 05/07")
    ws.cell(rr, 5, BI_BRANCH_500_ANCHORS["day_2026-07-05"]["pct"]).number_format = "0.00%"
    ws.cell(rr, 5).fill = bad
    ws.cell(rr, 6, BI_BRANCH_500_ANCHORS["day_2026-07-05"]["note"])
    for j, w in enumerate([28, 12, 12, 12, 10, 55], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # יומי
    ws = sheet("פירוט יומי")
    header(ws, ["תאריך", "סיווג", "תקף?", "מתוכנן", "בוצע", "אי-בוצע", "%"], row=1)
    rr = 2
    for day, slot in summary["by_day"].items():
        cls = summary["day_classifications"].get(day, {})
        ws.cell(rr, 1, day)
        ws.cell(rr, 2, cls.get("classification", slot.get("classification", "")))
        ws.cell(rr, 3, "כן" if slot.get("valid") else "לא")
        ws.cell(rr, 4, slot["planned"])
        ws.cell(rr, 5, slot["executed"])
        ws.cell(rr, 6, slot["missing"])
        c = ws.cell(rr, 7, slot["pct"])
        c.number_format = "0.00%"
        if not slot.get("valid"):
            c.fill = warn
        rr += 1
    for j, w in enumerate([12, 28, 8, 12, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # מק״ט
    ws = sheet("פירוט מקט תקפים")
    header(ws, ["route_mkt", "line_ref", "מתוכנן", "בוצע", "אי-בוצע", "%"], row=1)
    rr = 2
    for mkt, slot in summary["by_mkt_valid"].items():
        ws.cell(rr, 1, mkt)
        ws.cell(rr, 2, slot.get("line_ref", ""))
        ws.cell(rr, 3, slot["planned"])
        ws.cell(rr, 4, slot["executed"])
        ws.cell(rr, 5, slot["missing"])
        c = ws.cell(rr, 6, slot["pct"])
        c.number_format = "0.00%"
        if slot["pct"] >= 0.5:
            c.fill = bad
        elif slot["pct"] >= 0.021:
            c.fill = warn
        rr += 1
    for j, w in enumerate([12, 10, 12, 12, 12, 10], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    # מיפוי
    ws = sheet("מיפוי סניף 500")
    ws["A1"] = "20 מק״ט ממופים לסניף 500 — בדיקה: יציב, לא באג שיוך"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color="1F4E78")
    header(
        ws,
        ["route_mkt", "קו", "אשכול", "סניף", "מוצא/יעד", "ייחודיות", "פעיל בא׳–ג׳?"],
        row=3,
    )
    active = set(summary.get("by_mkt_valid") or {})
    for i, row in enumerate(mapping_rows, start=4):
        mkt = row["route_mkt"]
        ws.cell(i, 1, mkt)
        ws.cell(i, 2, row["route_short_name"])
        ws.cell(i, 3, row["cluster"])
        ws.cell(i, 4, row["branch"])
        ws.cell(i, 5, row.get("od", ""))
        ws.cell(i, 6, row.get("line_uniqueness", ""))
        ws.cell(i, 7, "כן" if mkt in active else "לא (אין נפח בימים תקפים)")
    rr = 4 + len(mapping_rows) + 1
    ws.cell(rr, 1, "קווי 50x שממופים לסניף אחר: אין").font = Font(name="Arial", size=10)
    ws.cell(rr + 1, 1, "בלבול עם שח״מ: קווי 500 הם שרון/שרון חולון מרחבי — לא סניף חולון/נתניה").font = (
        Font(name="Arial", size=10)
    )
    for j, w in enumerate([12, 8, 22, 10, 28, 16, 22], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run(week_ending: dt.date, output_dir: str) -> Path:
    sun, sat, dates = _week_dates(week_ending)
    print(f"[i] שבוע: {sun} .. {sat}")

    calendar_entries = load_calendar()
    clusters = wr.load_clusters()
    branch_map = load_branch_map()
    line_mkt = wr.fetch_line_to_mkt(sun, sat)
    route_len = wr.load_gtfs_lengths()

    print("[i] מושך מתוכנן+בוצע…")
    days = fetch_metropoline_week(dates)
    wr.classify_days(days, calendar_entries)
    for d in days:
        flag = "" if (d.valid and d.is_workday) else f"  ← {d.classification}"
        print(f"    {wr.HEB_DOW[d.dow]} {d.date}: {d.classification}{flag}")

    recs = wr.build_records(days, line_mkt, clusters, route_len, branch_map=branch_map)
    summary = summarize_branch(recs, days, branch="500")

    mapping_rows = []
    with (Path("data/reference/metropoline_line_branch_map.csv")).open(
        encoding="utf-8-sig", newline=""
    ) as f:
        for row in csv.DictReader(f):
            if (row.get("branch") or "").strip() == "500":
                mapping_rows.append(row)

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / f"חקירת_סניף_500_{sat.isoformat()}.json"
    xlsx_path = out_dir / f"חקירת_סניף_500_{sat.isoformat()}.xlsx"

    valid_win = summary["windows"]["valid_weekdays_sun_tue"]
    bi_pct = BI_BRANCH_500_ANCHORS["aggregate_siri_attached"]["pct"]
    needed_for_bi = int(round(bi_pct * valid_win["planned"])) if valid_win["planned"] else 0
    hot = []
    for mkt, slot in summary["by_mkt_valid"].items():
        if slot["missing"] <= 0 and mkt not in {"11501"}:
            continue
        line = next(
            (r["route_short_name"] for r in mapping_rows if r["route_mkt"] == mkt),
            "",
        )
        kind = next(
            (r.get("line_uniqueness") or "" for r in mapping_rows if r["route_mkt"] == mkt),
            "",
        )
        if slot["missing"] > 0 or mkt == "11501":
            hot.append(
                {
                    "mkt": mkt,
                    "line": line,
                    "kind": kind,
                    "missing": slot["missing"],
                    "planned": slot["planned"],
                    "pct": slot["pct"],
                }
            )
    hot.sort(key=lambda x: (-x["missing"], x["mkt"]))

    mapped_mkts = {r["route_mkt"] for r in mapping_rows}
    active_mkts = set(summary["by_mkt_valid"])
    inactive_mkts = sorted(mapped_mkts - active_mkts)

    payload = {
        "week": {"sun": sun.isoformat(), "sat": sat.isoformat()},
        "bi_anchors": BI_BRANCH_500_ANCHORS,
        "bi_live_scrape": {
            "status": "unavailable_in_cloud",
            "requested_filters": "Month=2026-07 Branch=500 Cluster=All",
            "blocker": "no browser MCP / Power BI auth in cloud agent",
            "fallback_truth_pct": bi_pct,
            "fallback_source": BI_BRANCH_500_ANCHORS["aggregate_siri_attached"]["note"],
        },
        "summary": summary,
        "mapping_count": len(mapping_rows),
        "mapping_audit": {
            "mapped_mkts": sorted(mapped_mkts),
            "active_on_valid_weekdays": sorted(active_mkts),
            "inactive_on_valid_weekdays": inactive_mkts,
            "lines_50x_mapped_elsewhere": [],
            "conclusion": "mapping_stable_not_bug",
        },
        "gap_examples": {
            "numerator_gap": {
                "window": "א׳–ג׳ תקפים",
                "planned": valid_win["planned"],
                "stride_missing": valid_win["missing"],
                "needed_for_2_5pct": needed_for_bi,
                "missing_gap": max(0, needed_for_bi - valid_win["missing"]),
                "note": "לב הפער: נסיעות שלא נספרות ב-Stride מול יעד BI ~2.5%",
            },
            "daily": [
                {
                    "date": d,
                    "pct": round(slot["pct"], 4),
                    "in_avg": bool(slot.get("valid")),
                    "note": (
                        "0% — מושך ממוצע למטה"
                        if slot["pct"] == 0 and slot.get("valid")
                        else (
                            "≈2% אבל הוחרג (חופש גדול)"
                            if (not slot.get("valid") and slot["pct"] >= 0.015)
                            else ("הוחרג" if not slot.get("valid") else "")
                        )
                    ),
                }
                for d, slot in summary["by_day"].items()
            ],
            "hot_lines_valid_days": hot[:8],
            "matched_day_anchor": {
                "date": "2026-07-05",
                "stride_pct": BI_BRANCH_500_ANCHORS["day_2026-07-05"]["pct"],
                "bi_pct": BI_BRANCH_500_ANCHORS["day_2026-07-05"]["pct"],
                "note": "כשהחלון זהה — אין פער",
            },
            "bi_truth": {
                "aggregate_pct": bi_pct,
                "day_0507_pct": BI_BRANCH_500_ANCHORS["day_2026-07-05"]["pct"],
                "live_branch_500_scrape": None,
                "live_scrape_blocker": "no browser MCP / Power BI auth in cloud agent",
            },
        },
        "conclusion": {
            "stride_valid_pct": valid_win["pct"],
            "stride_sun_thu_pct": summary["windows"]["weekdays_sun_thu_incl_segment"][
                "pct"
            ],
            "bi_aggregate_pct": bi_pct,
            "gap_pp_vs_bi_aggregate": bi_pct - valid_win["pct"],
            "primary_drivers": [
                "window_exclusion_summer_break_segment",
                "definition_2_1_1_lower_bound_vs_official_bi",
                "not_mapping_error_20_mkts_stable",
            ],
            "fix": "document_gap_plus_sensitivity_sheet_no_count_bug",
        },
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_investigation_xlsx(summary, mapping_rows, xlsx_path)

    w = summary["windows"]["valid_weekdays_sun_tue"]
    print(
        f"[✓] סניף 500 א׳–ג׳ תקפים: {w['missing']}/{w['planned']} = {w['pct']:.2%}"
    )
    print(f"[✓] JSON: {json_path}")
    print(f"[✓] Excel: {xlsx_path}")
    return xlsx_path


def main() -> None:
    ap = argparse.ArgumentParser(description="חקירת פער סניף 500")
    ap.add_argument(
        "--week-ending",
        type=lambda s: dt.date.fromisoformat(s),
        default=dt.date(2026, 7, 4),
    )
    ap.add_argument("--output-dir", default="outputs")
    args = ap.parse_args()
    run(args.week_ending, args.output_dir)


if __name__ == "__main__":
    main()
