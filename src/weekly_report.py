"""
דוח אי-ביצוע שבועי לתחבורה ציבורית בישראל — מרובה מפעילים.
מודד את רכיב אי-היציאה (2.1.1) בלבד = חסם תחתון לשיעור אי-הביצוע הרשמי.
הרצה:
  python -m src.weekly_report --week-ending 2026-06-27 --output-dir outputs
אם לא מצוין week-ending: השבוע שהסתיים אתמול (ראשון..שבת).
"""
from __future__ import annotations
import argparse, datetime as dt, math, statistics, io, zipfile, csv, collections
from concurrent.futures import ThreadPoolExecutor
import requests

from src.exclusions_calendar import (
    DateClassification,
    classify_date,
    load_calendar,
)

API = "https://open-bus-stride-api.hasadna.org.il"
GTFS_ZIP = "https://gtfs.mot.gov.il/gtfsfiles/israel-public-transportation.zip"
CLUSTER_ZIP = "https://gtfs.mot.gov.il/gtfsfiles/ClusterToLine.zip"
import os
def _zip_bytes(url, cache_name, timeout):
  os.makedirs("data", exist_ok=True)
  local = os.path.join("data", cache_name)
  if os.path.exists(local):
    with open(local, "rb") as fh:
      return fh.read()
  data = SESSION.get(url, timeout=timeout, headers=ZIP_HEADERS).content
  with open(local, "wb") as fh:
    fh.write(data)
  return data


OPERATORS = {  # operator_ref -> שם
    3: "אגד", 5: "דן", 15: "מטרופולין", 16: "סופרבוס",
    18: "קווים", 25: "אלקטרה אפיקים", 34: "תנופה",
}

# ספים (definitions.md / נספח כ"ו)
THRESH_SERVICE = 0.021      # מעל 2.1% = חריגה מרמת שירות
THRESH_FUNDAMENTAL = 0.025  # מעל 2.5% = הפרה יסודית
THRESH_PRECISION = 0.045    # אי-דיוק מעל 4.5% = חריגה

# מדרג פיצוי-לכל-הפרה (₪) לפי שיעור אי-הביצוע היומי באשכול — נספח כ"ו.
# הטבלה זהה במכרז 5/2021 (בקעת אונו) ובמכרז 04/2021 (השרון):
#   "שיעור יומי של כלל הנסיעות באשכול בהן נמצאו חריגות אי ביצוע ...
#    1.0% 1.5% 63 ... 3.5% 100% 173"
# מקור: meta.js: ono_37_5593C13A.bin עמ' 260-261; sharon_18_A4848144.bin עמ' 3.
GRADED_NONEXEC_TABLE: tuple[tuple[float, int], ...] = (
    (0.010, 0), (0.015, 63), (0.025, 93), (0.030, 118), (0.035, 143), (1.000, 173),
)


def tariff_from_table(rate: float, table: tuple = GRADED_NONEXEC_TABLE) -> int:
    """תעריף-להפרה לפי טבלת מדרגות: המדרגה הראשונה ש-rate ≤ גבולה העליון."""
    for upper, tariff in table:
        if rate <= upper:
            return tariff
    return table[-1][1]


def penalty_tariff(rate: float) -> int:
    """מדרג 0-173 ₪ להפרה (מכרזי 2021) — ר' GRADED_NONEXEC_TABLE."""
    return tariff_from_table(rate)

# בסיס P95 ליום חול (חושב 10.6.2026): (נסיעות, ק"מ)
P95 = {
    3:  (22475, 628021), 5:  (12112, 194884), 15: (12818, 328707),
    16: (11388, 219985), 18: (13778, 289334), 25: (8061, 155322),
    34: (2506, 95763),
}

# הנתיב האשכולי הישן של קנס קבוע 5,000 ₪ לאשכול-יום נשאר מת לצמיתות:
# נספח כ"ו מגדיר את הקנס הקבוע ברמת קו-יום (מכרז 5/2021 עמ' 261 §2.2), והוא
# ממומש כעת פר-משטר-מכרז ב-TENDER_REGIMES (ר' compute_penalties). שדה
# fixed_penalty ב-PenaltyRow נשאר 0 תמיד. הסף משמש לסימון over_threshold בלבד.
CLUSTER_DAY_FIXED_THRESHOLD = THRESH_PRECISION  # 4.5%

# "קו בתדירות נמוכה - קו אשר ביום נתון 100% מנסיעות הקו לכיוון מסוים, כמוגדר
#  ברישוי, הן בתדירות של 59 דקות ומעלה" (5/2021 עמ' 261; זהה ב-04/2021 עמ' 3)
LOW_FREQ_HEADWAY_MIN = 59

# חלון "אחרי חצות" מקומי (00:00-03:59) — יציאה בו מעידה על זיהום חלון ה-UTC
# היומי בזנבות של יום-השירות הקודם/הבא (ר' docstring של is_low_frequency).
NIGHT_WINDOW_END_HOUR = 4


def is_low_frequency(planned: int,
                     times: "list[dt.datetime] | None") -> "bool | None":
    """
    האם קו-יום הוא 'קו בתדירות נמוכה' (נספח כ"ו, מכרזי 5/2021 ו-04/2021):
    "קו אשר ביום נתון 100% מנסיעות הקו לכיוון מסוים, כמוגדר ברישוי, הן
    בתדירות של 59 דקות ומעלה".

    ההגדרה החוזית דורשת ש-100% מהנסיעות יעמדו במרווח ≥59 דק' — כלומר
    המרווח-העוקב ה*מינימלי* בין כל זוג נסיעות סמוכות חייב להיות ≥59 דק'.
    מרווח ממוצע אינו שקול: קו 06:00/06:15/06:30/18:00 (ממוצע 240 דק') הוא
    קו רגיל — המרווחים העוקבים של 15 דק' מפרים את ההגדרה.

    times = רשימה ממוינת של זמני scheduled_start_time מקומיים ייחודיים
    שנצפו ב-SIRI לקו-יום (fetch_executed).

    planned = סך הנסיעות המתוכננות לקו-יום, אחרי קאפ כפל-התכנון
    (apply_double_planning_cap). הסיווג עצמו מבוסס זמנים-נצפים בלבד;
    planned משמש רק לכלל הנסיעה-הבודדת ולשומר הזמנים-החלקיים שלהלן.
    אינטראקציה עם הקאפ: הקאפ רק מקטין את planned, ולכן מרכך את שומר
    הזמנים-החלקיים (len(times) < planned); בקצה, planned שהוקטן ל-1
    יסווג תדירות-נמוכה מכלל הנסיעה-הבודדת גם אם נצפו כמה זמנים.

    כללי הכרעה (בסדר):
    * planned <= 0 — None (לא-ניתן-לקביעה).
    * חציית-חצות: הזמנים נאספים מחלון UTC של יום קלנדרי מקומי, ולכן קו
      עם יציאות בחלון 00:00-03:59 מקומי עלול לכלול יציאות אחרי-חצות של
      יום-השירות הקודם ולהחסיר את זנב היום עצמו — המרווחים מזוהמים.
      אין כלל-שיוך אמין ליום-שירות (ולא ממציאים אחד), לכן יציאה כלשהי
      ב-00:00-03:59 ⇒ None. תיקון בכיוון השמרני בלבד: קו כזה לעולם לא
      מסווג תדירות-נמוכה.
    * planned == 1 — True: נסיעה מתוכננת אחת ביום, אין לה נסיעה עוקבת
      במרווח קצר מ-59 דק' — כל נסיעותיה 'נסיעות לא תדירות' (כלל מבוסס
      תכנון בלבד, אינו תלוי בזמנים הנצפים).
    * אין זמנים נצפים — None (planned >= 2 ללא זמני-יציאה).
    * len(times) < planned — זמנים חלקיים: המרווח-העוקב האמיתי יכול רק
      להיות קטן מהנצפה (נסיעות חסרות מפצלות מרווחים), ולכן סיווג
      'תדירות נמוכה' מזמנים חלקיים אינו בטוח ⇒ None.
    * אחרת: תדירות-נמוכה ⇔ המרווח-העוקב המינימלי ≥ 59 דק'.
      planned == 2: הפער היחיד בין שתי הנסיעות מכריע.

    מחזיר None כשלא ניתן לקבוע — הקו מטופל אז כקו רגיל והדבר מדווח
    (_regime_line_components), לא מנוחש.
    """
    if planned <= 0:
        return None
    if times and any(t.hour < NIGHT_WINDOW_END_HOUR for t in times):
        return None
    if planned == 1:
        return True
    if not times:
        return None
    if len(times) < planned:
        return None
    min_gap = min((t2 - t1).total_seconds() / 60.0
                  for t1, t2 in zip(times, times[1:]))
    return min_gap >= LOW_FREQ_HEADWAY_MIN


# ============================================================================
# משטרי-קנס פר-מכרז (נספח כ"ו) — מקור-האמת: חילוץ המכרזים (פסק נספח כ"ו 16.8).
# כל תעריף מלווה בציטוט והפניה למקור. עקרונות:
#   * רכיב שסכומו/הגדרתו לא נלכדו במקור — אינו מתומחר, מסומן ב-not_found
#     ומופיע בעמודת "פירוט רכיבים / לא נמצא במקור" בדוח. לעולם לא מנוחש.
#   * רכיב שמוגדר במקור אך אינו ניתן לחישוב מנתוני הדוח (ספירות אי-יציאה
#     בלבד) — מתועד ב-not_computable ואינו מתומחר. לכן החשיפה במשטרי 2021
#     (בקרה אלקטרונית) = חסם תחתון.
#   * משטרים ישנים (24/2015, 07/2014): הטריגר החוזי הוא "אירוע" מתועד
#     (דיווח פקח משה"ת / מוסמך מטעם המפקח / תלונת ציבור מוצדקת), לא בקרה
#     אלקטרונית על כלל הנסיעות — התמחור פר-נסיעת-SIRI הוא תקרת-חשיפה
#     תיאורטית, לא חסם תחתון. מתועד ברכיב trigger_note של המשטר ומופיע
#     בעמודת הפירוט בדוח.
#   * אשכול שאינו ב-TENDER_REGIMES אינו מתומחר בשקט: regime="לא-ממופה"
#     והשורה מסומנת בדוח.
# ============================================================================
_REGIME_ONO_ELAD = {
    "name": "מכרז 5/2021 — בקעת אונו (אונו-אלעד)",
    "graded": {
        "table": GRADED_NONEXEC_TABLE,
        "label": "מדרג אשכול-יומי 0-173 ₪/הפרה",
        "quote": "שיעור יומי של כלל הנסיעות באשכול בהן נמצאו חריגות אי ביצוע"
                 " ... 1.0% 1.5% 63 ... 3.5% 100% 173",
        "ref": "meta.js: ono_37_5593C13A.bin עמ' 260-261 (נספח פיצויים מוסכמים)",
    },
    "line_over_threshold": {
        "amount": 5000,
        "threshold": 0.045,
        "label": "§2.2 אי-ביצוע יומי בקו > 4.5% — 5,000 ₪ לקו-יום",
        "quote": "2.2 אי ביצוע יומי בקו שאינו מוגדר כקו בתדירות נמוכה עולה על"
                 " 4.5% - 5,000 ₪ לכל יום",
        "ref": "meta.js: ono_37_5593C13A.bin עמ' 261, §2.2",
    },
    "infrequent_ride": {
        "amount": 5000,
        "label": "§2.3 אי-ביצוע נסיעה לא תדירה (קו בתדירות נמוכה) — 5,000 ₪/נסיעה",
        "quote": "2.3 אי ביצוע של נסיעה לא תדירה – 5,000 ₪ לאירוע",
        "ref": "meta.js: ono_37_5593C13A.bin עמ' 261, §2.3 — הניתוב: קו בתדירות"
               " נמוכה מוחרג מ-§2.2 אך כל נסיעותיו 'לא תדירות' מעצם ההגדרה",
    },
    "not_found": [
        "מדרג האיחור §2.5.1 — הטבלה לא נלכדה בחילוץ (unresolved) — לא תומחר",
    ],
    "not_computable": [
        "§2.4 נסיעה אחרונה ביום (5,000 ₪) — זיהוי הנסיעה האחרונה דורש לוח-רישוי מלא",
        "§2.5.2 מדרג הקדמה (0-143 ₪) — אי-דיוק לא מדיד בנתוני המקור",
        "§2.9 תוספת תלונות ציבור (200/150 ₪) — אין נתוני תלונות",
        "§2.10 בקרה מדגמית — הדוח מבוסס בקרה אלקטרונית",
        "§§5-6 הצמדה למדד 2/2012 והכפלה בשנה האחרונה — לא הוחלו (סכומים נומינליים)",
        "נסיעות לא תדירות בודדות בקווים רגילים (פער-לו\"ז ≥59 דק' לנסיעה"
        " העוקבת) אינן מזוהות פר-נסיעה ואינן מתומחרות — תת-ספירה",
        "סיווגי אי-ביצוע 2.1.2-2.1.4 (איחור>20 דק', הקדמה≥10 דק', יציאה אחרי"
        " הנסיעה העוקבת) אינם מדידים — שיעור האשכול ומניין ההפרות הם חסם תחתון",
    ],
}

_REGIME_SHARON = {
    "name": "מכרז 04/2021 — השרון",
    "graded": {
        "table": GRADED_NONEXEC_TABLE,
        "label": "מדרג אשכול-יומי 0-173 ₪/הפרה",
        "quote": "שיעור יומי של כלל הנסיעות באשכול בהן נמצאו חריגות אי ביצוע"
                 " ... 1.0% 1.5% 63 ... 3.5% 100% 173",
        "ref": "meta.js: sharon_18_A4848144.bin עמ' 3 (העיגון הסעיפי המדויק"
               " לא חד-משמעי בחילוץ — כותרת הטבלה קושרת אותה לאי-ביצוע)",
    },
    "line_over_threshold": {
        "amount": 500,
        "threshold": 0.045,   # ביצוע < 95.5% ⇔ אי-ביצוע > 4.5%
        "label": "§2.1 ביצוע < 95.5% לתקופת-יום — 500 ₪ (חסם תחתון: תקופה אחת מתוך עד 3)",
        "quote": "2.1 ביצוע נמוך מ-95.5%, לתקופת יום בקו שאינו מוגדר כקו"
                 " בתדירות נמוכה - 500 ₪",
        "ref": "meta.js: sharon_18_A4848144.bin עמ' 3, §2.1; עד 3 קנסות ביום"
               " (תקופת-יום בנפרד): sharon_14 עמ' 48, שאלה 153",
        # אין בנתוני הדוח פיצול לתקופות-יום (המתוכנן ללא זמני-יציאה). יום
        # שביצועו הכולל < 95.5% מכיל בהכרח לפחות תקופת-יום אחת < 95.5%
        # (ממוצע משוקלל) — לכן תומחרה תקופה אחת בלבד = חסם תחתון, לא ניחוש.
    },
    "infrequent_ride": {
        "amount": 5000,
        "label": "§2.2 אי-ביצוע נסיעה לא תדירה (קו בתדירות נמוכה) — 5,000 ₪/נסיעה",
        "quote": "2.2 אי ביצוע של נסיעה לא תדירה – 5,000 ₪ לאירוע",
        "ref": "meta.js: sharon_18_A4848144.bin עמ' 3, §2.2; ניתוב קו בתדירות"
               " נמוכה: מוחרג מ-§2.1 אך כל נסיעותיו 'לא תדירות' (שאלת-הבהרה 152)",
    },
    "not_found": [
        "§2.3 אי-ביצוע נסיעה אחרונה ביום — הסעיף קיים (sharon_14 עמ' 44-45,"
        " שאלה 145) אך סכומו לא נלכד במקור — לא תומחר",
    ],
    "not_computable": [
        "מדרג איחור §2.4.1 (0-121 ₪) ומדרג הקדמה §2.4.2 (0-143 ₪) — אי-דיוק לא מדיד",
        "פיצוי מוגדל 130K/43K ₪ לכל 0.1% — מנגנון לתקופת-בקרה (סעיף 2.25), לא שבועי",
        "§2.8 תוספת תלונות ציבור; §2.10 בקרה מדגמית — אין נתונים",
        "§§5-6 הצמדה למדד 2/2012 והכפלה בשנה האחרונה — לא הוחלו (סכומים נומינליים)",
        "נסיעות לא תדירות בודדות בקווים רגילים (פער-לו\"ז ≥59 דק' לנסיעה"
        " העוקבת) אינן מזוהות פר-נסיעה ואינן מתומחרות — תת-ספירה",
        "סיווגי אי-ביצוע 2.1.2-2.1.4 (איחור>20 דק', הקדמה≥10 דק', יציאה אחרי"
        " הנסיעה העוקבת) אינם מדידים — שיעור האשכול ומניין ההפרות הם חסם תחתון",
    ],
}

_REGIME_SHARON_HOLON = {
    "name": "מכרז 24/2015 — שרון-חולון מרחבי",
    # אין במכרז זה מנגנון תקופת-יום, כלל 4.5%, טבלת-מדרגות או מושג
    # 'קו בתדירות נמוכה' — חיפוש ממצה ב-meta.js על כל עמודי 24/2015.
    "flat_per_ride": {
        "amount": 2000,
        "label": "אי-ביצוע נסיעה — 2,000 ₪/אירוע (לפני מקדם הכפלה)",
        "quote": "אי ביצוע נסיעה 2,000 אירוע",
        "ref": "meta.js: נספח כו' 24/2015, עמ' 263, טבלת §2 'לוח זמנים'",
    },
    # הטריגר החוזי לקנס במשטר זה הוא "אירוע" מתועד — לא בקרה אלקטרונית על
    # כלל הנסיעות. התמחור פר-נסיעת-SIRI שלא בוצעה הוא לכן תקרת-חשיפה
    # תיאורטית (כאילו כל אי-יציאה תועדה כאירוע), לא חסם תחתון.
    "trigger_note": (
        "הטריגר החוזי במכרז 24/2015 הוא אירוע מתועד (דיווח פקח משרד"
        " התחבורה / מוסמך מטעם המפקח / תלונת ציבור מוצדקת), לא בקרה"
        " אלקטרונית על כלל הנסיעות — התמחור פר-נסיעת-SIRI הוא תקרת-חשיפה"
        " תיאורטית, לא חסם תחתון"
    ),
    "not_found": [
        "מקדם ההכפלה שנקב המפעיל (1-5, מספרים טבעיים — עמ' 266+55) לא נמצא —"
        " תומחר לפני מקדם (מקדם ≥ 1 לא הוחל)",
        "אי-ביצוע בקו הזנה לרכבת — הסכום משובש בחילוץ ('00522') — לא תומחר בנפרד",
    ],
    "not_computable": [
        "חריגה מלו\"ז (500 ₪; 750 ₪ בקו הזנה) — אי-דיוק לא מדיד",
        "הצמדה למדד 2/2012 — לא הוחלה (סכומים נומינליים)",
    ],
}

_REGIME_NEGEV = {
    "name": "מכרז 07/2014 — הנגב",
    # אין במכרז זה מנגנון תקופת-יום, כלל 4.5%, טבלת-מדרגות או מושג
    # 'קו בתדירות נמוכה' — חיפוש ממצה ב-meta.js על כל עמודי 07/2014.
    "flat_per_ride": {
        "amount": 2000,
        "label": "אי-ביצוע נסיעה — 2,000 ₪/אירוע (לפני מקדם הכפלה)",
        "quote": "אי ביצוע נסיעה 2,000 אירוע",
        "ref": "meta.js: נספח כו' – פיצויים מוסכמים (קנסות) 07/2014, עמ' 316",
    },
    # ר' ההערה המקבילה ב-24/2015: הטריגר החוזי הוא אירוע מתועד, לא בקרה
    # אלקטרונית — התמחור פר-נסיעה הוא תקרת-חשיפה תיאורטית, לא חסם תחתון.
    "trigger_note": (
        "הטריגר החוזי במכרז 07/2014 הוא אירוע מתועד (דיווח פקח משרד"
        " התחבורה / מוסמך מטעם המפקח / תלונת ציבור מוצדקת), לא בקרה"
        " אלקטרונית על כלל הנסיעות — התמחור פר-נסיעת-SIRI הוא תקרת-חשיפה"
        " תיאורטית, לא חסם תחתון"
    ),
    "not_found": [
        "מקדם ההכפלה שנקב המפעיל (§28.5.4.2, מספרים טבעיים) לא נמצא —"
        " תומחר לפני מקדם (מקדם ≥ 1 לא הוחל)",
    ],
    "not_computable": [
        "אי-ביצוע בקו הזנה לרכבת (2,500 ₪) — אין סיווג קווי-הזנה בנתוני הדוח",
        "חריגה מלו\"ז (500 ₪; 750 ₪ בקו הזנה) — אי-דיוק לא מדיד",
    ],
}

# מיפוי אשכול (ClusterName כפי שמופיע ב-ClusterToLine.zip) -> משטר-מכרז.
# המשטרים כאן הם משטרי מכרזי מטרופולין בלבד, ולכן התמחור לפיהם מותנה גם
# במפעיל (METROPOLINE_REF) — שומר-המפעיל נאכף בנקודת ה-lookup של
# compute_penalties; שורת מפעיל אחר אינה מתומחרת ומסומנת כמו לא-ממופה.
# הערה עובדתית: 'צפון הנגב' בקובץ הארצי הוא אשכול של המפעיל "דן בדרום"
# (130 שורות ב-ClusterToLine.zip), לא של מטרופולין — הוא אינו ממופה כאן.
# אשכול הנגב של מטרופולין מופיע בקובץ כ-'הנגב' (90 שורות).
TENDER_REGIMES: dict[str, dict] = {
    "בקעת אונו אלעד": _REGIME_ONO_ELAD,
    "שרון": _REGIME_SHARON,
    "שרון חולון מרחבי": _REGIME_SHARON_HOLON,
    "הנגב": _REGIME_NEGEV,
}

# משטרי TENDER_REGIMES חלים על מטרופולין בלבד (ר' הערת המיפוי לעיל).
METROPOLINE_REF = 15
assert METROPOLINE_REF in OPERATORS and OPERATORS[METROPOLINE_REF] == "מטרופולין"

UNMAPPED_REGIME_LABEL = "לא-ממופה"

# ימי חול בישראל (Python weekday(): Mon=0 .. Sun=6) — ראשון עד חמישי
WEEKDAY_WORK = frozenset({6, 0, 1, 2, 3})   # א, ב, ג, ד, ה
WEEKDAY_FRIDAY = 4                            # ו
WEEKDAY_SATURDAY = 5                          # ש
HEB_DOW = {6: "א", 0: "ב", 1: "ג", 2: "ד", 3: "ה", 4: "ו", 5: "ש"}

# ספי זיהוי ימים חריגים
REDUCED_PLAN_RATIO = 0.75    # מתוכנן < 75% מהחציון => יום מופחת
STRIKE_EXEC_RATIO = 0.90     # ביצוע < 90% מהרגיל => חריג ביצוע (שביתה/אירוע)
DOUBLE_PLAN_RATIO = 1.7      # מתוכנן >= פי 1.7 מהחציון => כפל תכנון GTFS
SIRI_FAULT_RATE = 0.50       # ביצוע/מתוכנן נמוך מ-50% מהרגיל אצל רוב המפעילים => חשד תקלת SIRI

SESSION = requests.Session()
SESSION.headers.update({"Accept": "application/json"})
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
_retry = Retry(total=5, connect=5, read=5, backoff_factor=1.5, status_forcelist=(429, 500, 502, 503, 504), allowed_methods=frozenset(["GET"]))
_adapter = HTTPAdapter(max_retries=_retry, pool_maxsize=16)
SESSION.mount("https://", _adapter)
SESSION.mount("http://", _adapter)

# שרת ה-GTFS של משרד התחבורה מחזיר דף HTML אם נשלח Accept: application/json —
# הורדות ה-zip חייבות לבטל את הכותרת הזו.
ZIP_HEADERS = {"Accept": "*/*"}


# ---------- עזר: תאריכים וחלון UTC ----------
def week_ending_saturday(today: dt.date) -> tuple[dt.date, dt.date]:
    """מחזיר (ראשון, שבת) של השבוע שהסתיים אתמול. today=ראשון -> ראשון שעבר..שבת אמש."""
    # אתמול
    yest = today - dt.timedelta(days=1)
    # מצא את השבת האחרונה <= אתמול (פייתון: Sunday=6)
    # נרצה טווח ראשון..שבת. אם today ראשון, אתמול שבת.
    sat = yest
    while sat.weekday() != 5:  # 5 = Saturday
        sat -= dt.timedelta(days=1)
    sun = sat - dt.timedelta(days=6)
    return sun, sat

def is_idt(d: dt.date) -> bool:
    """קיץ ישראל (DST) — אפריל עד אוקטובר בקירוב. 6/2026 = קיץ (UTC+3)."""
    return 3 < d.month < 11 or (d.month == 3) or (d.month == 10 and d.day < 25)

def utc_window(d: dt.date) -> tuple[str, str]:
    """חלון UTC ליום מקומי. קיץ UTC+3: 21:00Z יום קודם .. 20:59:59Z. חורף UTC+2: 22:00Z."""
    off = 3 if is_idt(d) else 2
    start = dt.datetime(d.year, d.month, d.day) - dt.timedelta(hours=off)
    end = start + dt.timedelta(days=1) - dt.timedelta(seconds=1)
    fmt = "%Y-%m-%dT%H:%M:%SZ"
    return start.strftime(fmt), end.strftime(fmt)


# ---------- 1. מתוכנן לכל קו-יום ----------
def fetch_planned(d: dt.date) -> dict[tuple[int, str], int]:
    """GET /gtfs_rides_agg/group_by — קריאה אחת ליום. מפתח (operator_ref, line_ref) -> total_planned_rides."""
    ds = d.isoformat()
    r = SESSION.get(f"{API}/gtfs_rides_agg/group_by", params={
        "date_from": ds, "date_to": ds, "group_by": "operator_ref,line_ref",
    }, timeout=120)
    r.raise_for_status()
    out = {}
    for row in r.json():
        op = row.get("operator_ref"); line = str(row.get("line_ref"))
        if op in OPERATORS:
            out[(op, line)] = out.get((op, line), 0) + int(row.get("total_planned_rides") or 0)
    return out


# ---------- 2. בוצע (SIRI) לכל מפעיל-יום + 3. דה-דופ ----------
def fetch_executed(op: int, d: dt.date) -> tuple[dict[str, int], dict[str, list]]:
    """
    משיכת siri_rides במנות, דה-דופ לפי journey_ref.
    מחזיר (line_ref -> נסיעות שבוצעו,
           line_ref -> רשימה ממוינת של *כל* זמני scheduled_start_time
           המקומיים הייחודיים שנצפו לקו).
    רשימת הזמנים משמשת את is_low_frequency לסיווג 'קו בתדירות נמוכה' לפי
    המרווח-העוקב המינימלי (זמנים כפולים של אותו קו מאוחדים — set; זמן
    כפול אינו מרווח נצפה).
    """
    t_from, t_to = utc_window(d)
    off = 3 if is_idt(d) else 2
    seen_nonzero = set()        # journey_ref שלא מסתיים ב-0 — נספר פעם אחת
    counts = collections.Counter()  # line_ref -> count
    times: dict[str, set] = collections.defaultdict(set)  # line_ref -> {מקומי}
    offset, limit = 0, 10000
    while True:
        r = SESSION.get(f"{API}/siri_rides/list", params={
            "siri_route__operator_refs": op,
            "scheduled_start_time_from": t_from,
            "scheduled_start_time_to": t_to,
            "limit": limit, "offset": offset,
        }, timeout=180)
        r.raise_for_status()
        rows = r.json()
        if not rows:
            break
        for row in rows:
            jr = str(row.get("journey_ref") or "")
            line = str(row.get("siri_route__line_ref"))
            ts = row.get("scheduled_start_time")
            if ts:
                try:
                    t = dt.datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
                except ValueError:
                    t = None
                if t is not None:
                    if t.tzinfo is None:
                        t = t.replace(tzinfo=dt.timezone.utc)
                    local = (t.astimezone(dt.timezone.utc)
                             + dt.timedelta(hours=off)).replace(tzinfo=None)
                    times[line].add(local)
            if jr.endswith("-0"):
                counts[line] += 1                 # כל שורה נספרת
            else:
                if jr in seen_nonzero:
                    continue
                seen_nonzero.add(jr)
                counts[line] += 1                 # פעם אחת בלבד
        if len(rows) < limit:
            break
        offset += limit
    return dict(counts), {line: sorted(ts) for line, ts in times.items()}


# ---------- 4. הצלבה ברמת קו-יום ----------
def reconcile_day(planned: dict, executed_by_op: dict) -> dict:
    """
    לכל (op, line): בוצע = min(SIRI, מתוכנן). אי-ביצוע = מתוכנן - בוצע.
    מחזיר אגרגציה לפי מפעיל: {op: {'planned':..,'executed':..}}
    """
    agg = {op: {"planned": 0, "executed": 0} for op in OPERATORS}
    for (op, line), p in planned.items():
        ex = executed_by_op.get(op, {}).get(line, 0)
        done = min(ex, p)
        agg[op]["planned"] += p
        agg[op]["executed"] += done
    return agg


# ---------- 6. ק"מ מ-GTFS ----------
def haversine(a, b) -> float:
    R = 6371.0088
    lat1, lon1, lat2, lon2 = map(math.radians, [a[0], a[1], b[0], b[1]])
    dlat, dlon = lat2 - lat1, lon2 - lon1
    h = math.sin(dlat/2)**2 + math.cos(lat1)*math.cos(lat2)*math.sin(dlon/2)**2
    return 2 * R * math.asin(math.sqrt(h))

def load_gtfs_lengths() -> dict[str, float]:
    """
    מוריד את ה-GTFS הארצי, מחשב אורך כל route_id (ק"מ) לפי ה-shape השכיח בטריפים שלו.
    הצלבה: route_id == line_ref. מחזיר line_ref(str) -> אורך_ק"מ.
    """
    z = zipfile.ZipFile(io.BytesIO(_zip_bytes(GTFS_ZIP, "israel-public-transportation.zip", 600)))

    def read(name):
          # קבצי ה-GTFS של משרד התחבורה כוללים BOM — utf-8-sig מסיר אותו,
          # אחרת שם העמודה הראשונה מקבל תחילית '﻿' (KeyError על shape_id/route_id).
          with z.open(name) as f:
                return list(csv.DictReader(io.TextIOWrapper(f, "utf-8-sig")))

    # shapes: shape_id -> [(lat,lon)] ממוין לפי shape_pt_sequence
    shape_pts = collections.defaultdict(list)
    for row in read("shapes.txt"):
        shape_pts[row["shape_id"]].append(
            (int(row["shape_pt_sequence"]), float(row["shape_pt_lat"]), float(row["shape_pt_lon"]))
        )
    shape_len = {}
    for sid, pts in shape_pts.items():
        pts.sort()
        total = sum(haversine((pts[i][1], pts[i][2]), (pts[i+1][1], pts[i+1][2]))
                    for i in range(len(pts)-1))
        shape_len[sid] = total

    # trips: route_id -> Counter(shape_id) כדי לבחור shape שכיח
    route_shapes = collections.defaultdict(collections.Counter)
    for row in read("trips.txt"):
        if row.get("shape_id"):
            route_shapes[row["route_id"]][row["shape_id"]] += 1

    route_len = {}
    for rid, ctr in route_shapes.items():
        common_shape = ctr.most_common(1)[0][0]
        route_len[rid] = shape_len.get(common_shape, 0.0)
    return route_len  # ק"מ לנסיעה אחת בקו


# ---------- 5. זיהוי: קו -> אשכול ו-קו -> route_mkt ----------
def _strip_zeros(code: str) -> str:
    """מסיר אפסים מובילים לצורך השוואת OfficeLineId == route_mkt."""
    s = str(code or "").strip().lstrip("0")
    return s or "0"


def _parse_cluster_date(s: str) -> "dt.date | None":
    """תאריך מ-ClusterToLine (DD/MM/YYYY; ISO כגיבוי). None אם לא פריק."""
    s = (s or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def load_clusters(today: "dt.date | None" = None) -> dict[str, str]:
    """
    מוריד את ClusterToLine.zip ומחזיר מיפוי route_mkt(ללא אפסים מובילים) -> שם אשכול.
    OfficeLineId הוא ה-route_mkt.

    סינון תוקף (ToDate): רשומה שפג תוקפה — ToDate לפני today — נזרקת,
    כך שאינה תופסת ב-setdefault את מקומה של הרשומה הפעילה של אותו מק"ט.
    ToDate בפורמט DD/MM/YYYY (רשומות פעילות נושאות 01/01/2200 בקובץ
    הארצי). רשומה עם ToDate חסר או בלתי-פריק נשמרת — אין זריקה שקטה
    בגלל שיבוש פורמט.
    """
    if today is None:
        today = dt.date.today()
    z = zipfile.ZipFile(io.BytesIO(_zip_bytes(CLUSTER_ZIP, "ClusterToLine.zip", 300)))
    name = z.namelist()[0]
    out: dict[str, str] = {}
    with z.open(name) as f:
        for row in csv.DictReader(io.TextIOWrapper(f, "utf-8-sig")):
            to_date = _parse_cluster_date(row.get("ToDate"))
            if to_date is not None and to_date < today:
                continue                       # פג תוקף — נזרק
            mkt = _strip_zeros(row.get("OfficeLineId"))
            cluster = (row.get("ClusterName") or "").strip()
            if mkt and cluster:
                out.setdefault(mkt, cluster)   # ההתאמה הראשונה (קלאסטר ראשי) מנצחת
    return out


def fetch_line_to_mkt(dates: "dt.date | list[dt.date]") -> dict[tuple[int, str], str]:
    """
    מיפוי (operator_ref, line_ref) -> route_mkt דרך /gtfs_routes/list, רק למפעילים שלנו.

    חשוב: אין להסתפק ב-snapshot של יום שבת. בשבת פועל רק חלק קטן מהקווים, ולכן
    מיפוי שנבנה משבת בלבד מחמיץ כמחצית מהקווים (מטרופולין: 306 מתוך 614), וכל
    הקווים החסרים נופלים לאשכול "לא משויך" — מה שמעוות את הפילוח לפי אשכול ואת
    חישוב הקנסות המדורגים. מקבל תאריך בודד או רשימה; המיפוי הוא איחוד, כאשר
    תאריך מאוחר יותר ברשימה גובר. מומלץ לקרוא עם [שבת, יום-חול].
    """
    if isinstance(dates, dt.date):
        dates = [dates]
    out: dict[tuple[int, str], str] = {}
    for d in dates:
      for op in OPERATORS:
        offset, limit = 0, 5000
        while True:
            r = SESSION.get(f"{API}/gtfs_routes/list", params={
                "date_from": d.isoformat(), "date_to": d.isoformat(),
                "operator_refs": op, "limit": limit, "offset": offset,
            }, timeout=120)
            r.raise_for_status()
            rows = r.json()
            if not rows:
                break
            for row in rows:
                out[(op, str(row.get("line_ref")))] = _strip_zeros(row.get("route_mkt"))
            if len(rows) < limit:
                break
            offset += limit
    return out


def cluster_for(op: int, line: str, line_mkt: dict, clusters: dict) -> str:
    """אשכול של קו, או 'לא משויך' אם אין התאמה."""
    mkt = line_mkt.get((op, line))
    if not mkt:
        return "לא משויך"
    return clusters.get(mkt, "לא משויך")


# ---------- פרופיל שעתי לאימות תקלת SIRI ----------
def fetch_hourly_profile(op: int, d: dt.date) -> dict[int, int]:
    """מספר נסיעות SIRI לפי שעת scheduled_start_time מקומית. לזיהוי 'נפילת מצוק' לאפס."""
    t_from, t_to = utc_window(d)
    off = 3 if is_idt(d) else 2
    hours = collections.Counter()
    offset, limit = 0, 10000
    while True:
        r = SESSION.get(f"{API}/siri_rides/list", params={
            "siri_route__operator_refs": op,
            "scheduled_start_time_from": t_from, "scheduled_start_time_to": t_to,
            "limit": limit, "offset": offset,
        }, timeout=180)
        r.raise_for_status()
        rows = r.json()
        if not rows:
            break
        for row in rows:
            ts = row.get("scheduled_start_time")
            if not ts:
                continue
            # ISO עם offset; נמיר לשעה מקומית
            t = dt.datetime.fromisoformat(ts.replace("Z", "+00:00"))
            local = t + dt.timedelta(hours=off) if t.tzinfo is None else \
                    (t.astimezone(dt.timezone.utc) + dt.timedelta(hours=off))
            hours[local.hour] += 1
        if len(rows) < limit:
            break
        offset += limit
    return dict(hours)


def detect_cliff(profile: dict[int, int], active_from: int = 5, active_to: int = 22) -> int | None:
    """
    מזהה 'נפילת מצוק' לאפס: שעה שממנה ואילך (בתוך טווח פעילות) אין כמעט נסיעות,
    בעוד שלפניה היו. מחזיר את שעת הנפילה, או None אם אין.
    """
    active = [h for h in range(active_from, active_to + 1)]
    if sum(profile.get(h, 0) for h in active) == 0:
        return active_from  # אין נסיעות כלל בטווח הפעיל
    for h in active:
        before = sum(profile.get(x, 0) for x in range(active_from, h))
        after = sum(profile.get(x, 0) for x in range(h, active_to + 1))
        if before >= 50 and after == 0:
            return h
    return None


# ---------- מבני נתונים ----------
class DayData:
    """נתוני יום בודד: מתוכנן/בוצע ברמת (op,line), סיווג, ודגלים."""
    def __init__(self, date_: dt.date):
        self.date = date_
        self.dow = date_.weekday()
        self.is_workday = self.dow in WEEKDAY_WORK
        self.planned_raw: dict[tuple[int, str], int] = {}   # לפני קאפ כפל-תכנון
        self.planned: dict[tuple[int, str], int] = {}       # אחרי קאפ
        self.executed: dict[int, dict[str, int]] = {}       # op -> {line: בוצע}
        # (op, line) -> רשימה ממוינת של זמני-יציאה מתוכננים מקומיים ייחודיים
        # שנצפו ב-SIRI — לסיווג תדירות (is_low_frequency)
        self.sched_times: dict[tuple[int, str], list] = {}
        self.calendar: DateClassification | None = None     # סיווג מלוח החגים
        self.classification = "תקין"                         # תקין/מופחת/חריג-ביצוע/לא-תקף/החרגה
        self.reason = ""
        self.strike_ops: set[int] = set()                   # מפעילים עם חריג ביצוע נקודתי
        self.valid = True                                   # נכלל בממוצעים ובחשיפה
        self.segment: "str | None" = None                   # שם תקופת ה-segment, אם יש
        self.percentages = True                             # האם להציג אחוזים בכלל

    def exec_for(self, op: int, line: str) -> int:
        return self.executed.get(op, {}).get(line, 0)


def fetch_day(d: dt.date, line_mkt: dict) -> DayData:
    """מושך מתוכנן + בוצע (כל המפעילים במקביל) ליום אחד."""
    day = DayData(d)
    day.planned_raw = fetch_planned(d)
    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = {op: ex.submit(fetch_executed, op, d) for op in OPERATORS}
        for op, f in futs.items():
            counts, times = f.result()
            day.executed[op] = counts
            for line, tlist in times.items():
                day.sched_times[(op, line)] = tlist
    return day


# ---------- 5. זיהוי ימים חריגים ----------
def _op_planned_total(day: DayData, source: str = "planned_raw") -> dict[int, int]:
    """סך מתוכנן לכל מפעיל ביום."""
    src = getattr(day, source)
    tot = collections.Counter()
    for (op, _line), p in src.items():
        tot[op] += p
    return dict(tot)


def _op_executed_total(day: DayData) -> dict[int, int]:
    tot = {}
    for op in OPERATORS:
        tot[op] = sum(day.executed.get(op, {}).values())
    return tot


def apply_double_planning_cap(workdays: list[DayData]) -> int:
    """
    כפל תכנון GTFS: לכל קו, אם מתוכנן ביום >= פי 1.7 מחציון ימי החול -> הצב חציון.
    מיושם אחיד על כל ימי החול. מעדכן day.planned (מאותחל מ-planned_raw).
    מחזיר מספר התאמות שבוצעו.
    """
    # אסוף לכל (op,line) את רשימת המתוכנן על פני ימי החול
    per_line: dict[tuple[int, str], list[int]] = collections.defaultdict(list)
    for day in workdays:
        for key, p in day.planned_raw.items():
            per_line[key].append(p)
    medians = {key: statistics.median(vals) for key, vals in per_line.items() if vals}

    fixes = 0
    for day in workdays:
        day.planned = dict(day.planned_raw)
        for key, p in day.planned_raw.items():
            med = medians.get(key, 0)
            if med > 0 and p >= DOUBLE_PLAN_RATIO * med:
                day.planned[key] = int(round(med))
                fixes += 1
    # ימים שאינם ימי-חול: planned = planned_raw ללא קאפ (דפוס שונה ממילא)
    return fixes


def classify_days(days: list[DayData], calendar_entries) -> None:
    """
    מסווג כל יום: לוח חגים -> מופחת -> חריג-ביצוע (פר-מפעיל) -> תקלת SIRI (חוצת-מפעילים).
    מעדכן day.classification/reason/valid/percentages/strike_ops.
    """
    workdays = [d for d in days if d.is_workday]

    # קאפ כפל-תכנון (גם לימים שאינם חול — planned=planned_raw)
    apply_double_planning_cap(workdays)
    for d in days:
        if not d.planned:
            d.planned = dict(d.planned_raw)

    # חציון מתוכנן/בוצע פר-מפעיל על ימי החול (בסיס להשוואות)
    plan_by_op: dict[int, list[int]] = collections.defaultdict(list)
    exec_by_op: dict[int, list[int]] = collections.defaultdict(list)
    for d in workdays:
        pt, et = _op_planned_total(d, "planned"), _op_executed_total(d)
        for op in OPERATORS:
            plan_by_op[op].append(pt.get(op, 0))
            exec_by_op[op].append(et.get(op, 0))
    plan_med = {op: statistics.median(v) if v else 0 for op, v in plan_by_op.items()}
    exec_med = {op: statistics.median(v) if v else 0 for op, v in exec_by_op.items()}

    n_ops = len(OPERATORS)
    majority = n_ops // 2 + 1

    for d in days:
        # 1) לוח חגים ארצי
        d.calendar = classify_date(d.date, calendar_entries, national_only=True)
        if d.calendar.treatment == "drop":
            d.classification = "החרגה (לוח)"
            d.reason = f"לוח החרגות: {d.calendar.name}"
            d.valid = False
            continue
        if d.calendar.treatment == "segment":
            # segment = "דפוס אמיתי וחוזר אך שונה → מנותח בנפרד, לא נזרק"
            # (data/reference/exclusions_2026/README.md). היום נשאר תקף ונכלל
            # בקנסות ובממוצעים; בסיס ההשוואה הוא קבוצת ה-segment עצמה.
            # פסילת היום כאן ביטלה את כל יולי-אוגוסט ("חופש גדול", 1.7-31.8
            # ארצי) ואיפסה את חשיפת הקנסות לשישית מהשנה.
            d.classification = "תקין (segment)"
            d.reason = f"לוח החרגות: {d.calendar.name} — דפוס שונה, מושווה מול תקופת ה-segment"
            d.segment = d.calendar.name

        if not d.is_workday:
            # שישי/שבת: דפוס בסיס שונה — מוצג בנפרד, לא בממוצעי יום-חול
            if d.classification == "תקין":
                d.classification = "שישי/שבת (דפוס נפרד)"
                d.reason = f"יום {HEB_DOW[d.dow]} — דפוס בסיס שונה"
            d.valid = False
            continue

    # רק על ימי חול שעדיין 'תקין' נריץ זיהוי מופחת/שביתה/SIRI
    candidate = [d for d in workdays if d.classification == "תקין"]

    # 2) יום מופחת מתוכנן (חג/ערב חג/צום שלא תויג בלוח) — מול תקלת נתוני-תכנון.
    #    יום מופחת אמיתי: גם המתוכנן וגם הביצוע נמוכים. פער נתוני-תכנון: המתוכנן
    #    כמעט אפס אך הביצוע תקין (gtfs_rides_agg לא אוכלס לאותו תאריך) — לא אי-ביצוע.
    for d in candidate:
        pt, et = _op_planned_total(d, "planned"), _op_executed_total(d)
        reduced = sum(1 for op in OPERATORS
                      if plan_med.get(op, 0) > 0 and pt.get(op, 0) < REDUCED_PLAN_RATIO * plan_med[op])
        if reduced < majority:
            continue
        exec_normal = sum(1 for op in OPERATORS
                          if exec_med.get(op, 0) > 0 and et.get(op, 0) >= STRIKE_EXEC_RATIO * exec_med[op])
        if exec_normal >= majority:
            d.classification = "לא תקף (תכנון חסר)"
            d.reason = (f"מתוכנן < 75% מהחציון אצל {reduced}/{n_ops} מפעילים אך הביצוע תקין "
                        f"({exec_normal}/{n_ops}) — פער נתוני-תכנון ב-gtfs_rides_agg, לא אי-ביצוע")
            d.valid = False
            d.percentages = False
        else:
            d.classification = "מופחת (תכנון)"
            d.reason = f"מתוכנן ובוצע < מהחציון אצל רוב המפעילים ({reduced}/{n_ops}) — יום שירות מופחת"
            d.valid = False

    # 3) תקלת SIRI חוצת-מפעילים: ביצוע נמוך חריג אצל רוב המפעילים בו-זמנית
    for d in [x for x in candidate if x.classification == "תקין"]:
        et = _op_executed_total(d)
        low = 0
        for op in OPERATORS:
            if exec_med.get(op, 0) > 0 and et.get(op, 0) < SIRI_FAULT_RATE * exec_med[op]:
                low += 1
        if low >= majority:
            # אימות: פרופיל שעתי ל-3 מפעילים גדולים — נפילת מצוק לאפס אצל כולם
            sample_ops = sorted(OPERATORS, key=lambda o: -plan_med.get(o, 0))[:3]
            cliffs = {op: detect_cliff(fetch_hourly_profile(op, d.date)) for op in sample_ops}
            if all(c is not None for c in cliffs.values()):
                d.classification = "לא תקף (תקלת SIRI)"
                hrs = ", ".join(f"{HEB_DOW[d.dow]}:{OPERATORS[o]}≈{cliffs[o]}:00" for o in sample_ops)
                d.reason = (f"ביצוע נמוך אצל {low}/{n_ops} מפעילים + נפילת מצוק שעתית "
                            f"({hrs}) — פער נתונים, לא אי-ביצוע")
                d.valid = False
                d.percentages = False
                continue

    # 4) חריג ביצוע נקודתי (שביתה/אירוע פר-מפעיל): תכנון רגיל, ביצוע < 90% מהרגיל
    for d in workdays:
        if not d.valid:
            continue
        pt, et = _op_planned_total(d, "planned"), _op_executed_total(d)
        for op in OPERATORS:
            plan_ok = plan_med.get(op, 0) == 0 or pt.get(op, 0) >= REDUCED_PLAN_RATIO * plan_med[op]
            if plan_ok and exec_med.get(op, 0) > 0 and et.get(op, 0) < STRIKE_EXEC_RATIO * exec_med[op]:
                d.strike_ops.add(op)
        if d.strike_ops:
            names = ", ".join(OPERATORS[o] for o in sorted(d.strike_ops))
            note = f"חריג ביצוע (שביתה/אירוע): {names} — מוחרגים מהממוצע"
            d.reason = (d.reason + " | " + note) if d.reason else note


# ---------- 6+7. הצלבה ברמת (op, line, cluster, day) + אגרגציות ----------
class LineDayRecord:
    __slots__ = ("date", "dow", "op", "op_name", "line", "mkt", "cluster",
                 "planned", "executed", "nonexec", "km_per_ride",
                 "planned_km", "executed_km", "nonexec_km")

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw.get(k))


def build_records(days: list[DayData], line_mkt: dict, clusters: dict,
                  route_len: dict) -> list[LineDayRecord]:
    """רשומת (op,line,cluster,day) עם בוצע=min(SIRI,מתוכנן), אי-ביצוע, וק"מ."""
    recs: list[LineDayRecord] = []
    for d in days:
        for (op, line), p in d.planned.items():
            if op not in OPERATORS:
                continue
            ex = min(d.exec_for(op, line), p)
            kmr = route_len.get(line, 0.0)
            recs.append(LineDayRecord(
                date=d.date, dow=HEB_DOW[d.dow], op=op, op_name=OPERATORS[op],
                line=line, mkt=line_mkt.get((op, line), ""),
                cluster=cluster_for(op, line, line_mkt, clusters),
                planned=p, executed=ex, nonexec=p - ex, km_per_ride=round(kmr, 3),
                planned_km=p * kmr, executed_km=ex * kmr, nonexec_km=(p - ex) * kmr,
            ))
    return recs


def aggregate_operator(days: list[DayData], recs: list[LineDayRecord]) -> dict:
    """
    אגרגציה פר-מפעיל על ימי חול תקפים בלבד, תוך החרגת מפעילים בחריג-ביצוע נקודתי.
    מחזיר op -> {planned, executed, nonexec, planned_km, executed_km, nonexec_km, valid_days}
    """
    valid_dates = {d.date for d in days if d.valid and d.is_workday}
    strike = {(d.date, op) for d in days for op in d.strike_ops}
    out = {op: collections.Counter() for op in OPERATORS}
    vdays = {op: set() for op in OPERATORS}
    for r in recs:
        if r.date not in valid_dates or (r.date, r.op) in strike:
            continue
        c = out[r.op]
        c["planned"] += r.planned; c["executed"] += r.executed; c["nonexec"] += r.nonexec
        c["planned_km"] += r.planned_km; c["executed_km"] += r.executed_km
        c["nonexec_km"] += r.nonexec_km
        vdays[r.op].add(r.date)
    return {op: {**dict(out[op]), "valid_days": len(vdays[op])} for op in OPERATORS}


# ---------- 2. חשיפת קנסות (נספח כ"ו) — מנוע פר-משטר-מכרז ----------
class PenaltyRow:
    __slots__ = ("date", "dow", "op", "op_name", "cluster", "regime",
                 "planned", "executed", "nonexec", "rate", "tariff",
                 "exposure", "components", "fixed_total", "total",
                 "not_found", "over_threshold", "fixed_penalty")

    def __init__(self, **kw):
        for k in self.__slots__:
            setattr(self, k, kw.get(k))


def _regime_line_components(regime: dict, line_recs: list,
                            day: "DayData | None") -> tuple[list, list]:
    """
    רכיבי-הקנס ברמת קו-יום/נסיעה של משטר אחד, על רשומות קו של דלי
    (יום, מפעיל, אשכול) אחד. מחזיר (components, notes):
      components = [(תווית רכיב, סכום ₪)], notes = הערות-חישוב לשקיפות.

    * flat_per_ride (24/2015, 07/2014): סכום קבוע × כל נסיעה שלא בוצעה —
      ללא כלל 4.5%, ללא מדרג וללא מושג 'קו בתדירות נמוכה' במכרזים אלה.
    * infrequent_ride (מכרזי 2021): קו בתדירות נמוכה (is_low_frequency)
      מנותב לסעיף המחמיר — כל נסיעה שלא בוצעה בו × 5,000 ₪; הוא מוחרג
      מהקנס הקבוע של קו רגיל.
    * line_over_threshold (מכרזי 2021): קו רגיל שאי-הביצוע היומי שלו חצה
      את סף המקור (4.5% ⇔ ביצוע<95.5%) — קנס קבוע לקו-יום.
    * קו שתדירותו לא ניתנת לקביעה (is_low_frequency מחזיר None: אין
      זמני-יציאה, זמנים חלקיים, או יציאות בחלון 00:00-03:59) מטופל
      כקו רגיל ומדווח.
    """
    comps: list[tuple[str, int]] = []
    notes: list[str] = []

    flat = regime.get("flat_per_ride")
    if flat:
        n = sum(r.nonexec for r in line_recs)
        if n > 0:
            comps.append((f"{flat['label']} × {n}", n * flat["amount"]))

    lot = regime.get("line_over_threshold")
    infreq = regime.get("infrequent_ride")
    if not (lot or infreq):
        return comps, notes

    times_map = day.sched_times if day is not None else {}
    lf_rides = 0      # נסיעות שלא בוצעו בקווי תדירות-נמוכה
    over_lines = 0    # קווי-יום רגילים שחצו את הסף
    unknown = 0       # קווים שתדירותם לא ניתנת לקביעה
    for r in line_recs:
        if r.planned <= 0:
            continue
        lf = is_low_frequency(r.planned, times_map.get((r.op, r.line)))
        if lf is True:
            if infreq:
                lf_rides += r.nonexec
            continue      # קו בתדירות נמוכה מוחרג מהקנס הקבוע של קו רגיל
        if lf is None:
            unknown += 1  # מטופל כקו רגיל — נופל להמשך הבדיקה
        if lot and r.nonexec / r.planned > lot["threshold"]:
            over_lines += 1
    if infreq and lf_rides > 0:
        comps.append((f"{infreq['label']} × {lf_rides}", lf_rides * infreq["amount"]))
    if lot and over_lines > 0:
        comps.append((f"{lot['label']} × {over_lines} קווים", over_lines * lot["amount"]))
    if unknown:
        notes.append(f"{unknown} קווים שתדירותם לא ניתנת לקביעה "
                     "(זמני-יציאה חסרים/חלקיים או יציאות 00:00-03:59) — טופלו כקו רגיל")
    return comps, notes


def compute_penalties(days: list[DayData], recs: list[LineDayRecord]) -> list[PenaltyRow]:
    """
    מנוע-קנסות פר-משטר-מכרז (TENDER_REGIMES) לכל מפעיל×אשכול×יום-חול-תקף.

    לכל דלי (יום, מפעיל, אשכול):
      * המשטר נקבע לפי האשכול — אך רק לשורות מטרופולין (METROPOLINE_REF):
        משטרי TENDER_REGIMES הם משטרי מכרזי מטרופולין, ולכן שורת מפעיל
        אחר אינה מתומחרת לפיהם ומסומנת כמו לא-ממופה (עם נימוק משלה).
        אשכול שאינו ממופה לאף משטר אינו מתומחר בשקט:
        regime="לא-ממופה", הסכומים None (לא 0) והשורה מסומנת בדוח.
      * רכיב מדורג (מכרזי 2021 בלבד — רכיב אשכולי מפורש במקור): תעריף-להפרה
        לפי שיעור אי-הביצוע היומי באשכול; חשיפה מדורגת = אי-בוצע × תעריף.
      * רכיבי קו-יום/נסיעה: ר' _regime_line_components (קו בתדירות נמוכה
        מנותב לסעיף המחמיר; קו רגיל מעל הסף מקבל את הקנס הקבוע של המשטר;
        מכרזים ישנים — סכום קבוע לכל נסיעה, ללא כלל 4.5%).
      * רכיבים שסכומם לא נלכד במקור אינם מתומחרים ומסומנים ב-not_found.

    הנתיב האשכולי הישן של קנס קבוע 5,000 ₪ לאשכול-יום נשאר מת:
    fixed_penalty=0 תמיד. over_threshold ממשיך לסמן חציית 4.5% (הדגשה בלבד).
    מחריג ימים לא-תקפים ומפעילים בחריג-ביצוע נקודתי.
    """
    valid_dates = {d.date for d in days if d.valid and d.is_workday}
    strike = {(d.date, op) for d in days for op in d.strike_ops}
    day_by_date = {d.date: d for d in days}
    bucket: dict[tuple, list] = collections.defaultdict(list)
    dow = {d.date: HEB_DOW[d.dow] for d in days}
    for r in recs:
        if r.date not in valid_dates or (r.date, r.op) in strike:
            continue
        bucket[(r.date, r.op, r.cluster)].append(r)

    rows: list[PenaltyRow] = []
    for (date_, op, cluster), rlist in sorted(bucket.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2])):
        planned = sum(r.planned for r in rlist)
        if planned == 0:
            continue
        executed = sum(r.executed for r in rlist)
        nonexec = sum(r.nonexec for r in rlist)
        rate = nonexec / planned
        over = rate > CLUSTER_DAY_FIXED_THRESHOLD
        base = dict(date=date_, dow=dow.get(date_, ""), op=op, op_name=OPERATORS.get(op, str(op)),
                    cluster=cluster, planned=planned, executed=executed, nonexec=nonexec,
                    rate=rate, over_threshold=over, fixed_penalty=0)  # הנתיב הישן מת

        # שומר-מפעיל: משטרי TENDER_REGIMES הם משטרי מכרזי מטרופולין בלבד.
        regime = TENDER_REGIMES.get(cluster) if op == METROPOLINE_REF else None
        if regime is None:
            if op != METROPOLINE_REF:
                reason = (f"משטרי TENDER_REGIMES חלים על מטרופולין "
                          f"(operator_ref {METROPOLINE_REF}) בלבד — שורת "
                          f"{OPERATORS.get(op, op)} לא תומחרה")
            else:
                reason = "אשכול ללא משטר-מכרז ממופה — החשיפה לא תומחרה (אין לתמחר 0 בשקט)"
            rows.append(PenaltyRow(
                **base, regime=UNMAPPED_REGIME_LABEL,
                tariff=None, exposure=None, components=[], fixed_total=None, total=None,
                not_found=reason,
            ))
            continue

        tariff = exposure = None
        graded = regime.get("graded")
        if graded:
            tariff = tariff_from_table(rate, graded["table"])
            exposure = nonexec * tariff
        comps, notes = _regime_line_components(regime, rlist, day_by_date.get(date_))
        fixed_total = sum(a for _, a in comps)
        nf = list(regime.get("not_found", ())) + notes
        if regime.get("trigger_note"):
            nf.append(regime["trigger_note"])
        rows.append(PenaltyRow(
            **base, regime=regime["name"], tariff=tariff, exposure=exposure,
            components=comps, fixed_total=fixed_total,
            total=(exposure or 0) + fixed_total, not_found="; ".join(nf),
        ))
    return rows


def summarize_penalties_by_cluster(rows: list[PenaltyRow]) -> list[dict]:
    """
    סיכום חשיפה פר-(אשכול, משטר): מתוכנן/אי-בוצע, חשיפה מדורגת, רכיבי
    קו/נסיעה, סה"כ, ומניין שורות מפעיל-יום מעל 4.5%. המפתח כולל את המשטר
    כי אותו אשכול יכול להופיע גם ממופה (שורות מטרופולין) וגם "לא-ממופה"
    (שורות מפעיל אחר — שומר-המפעיל) — אסור לערבב אותם בשורת-סיכום אחת.
    שורות לא-ממופות מסומנות priced=False — סכומיהן אינם 0 אלא "לא תומחר".
    """
    agg: dict[tuple, dict] = {}
    for p in rows:
        s = agg.setdefault((p.cluster, p.regime), {
            "cluster": p.cluster, "regime": p.regime,
            "planned": 0, "nonexec": 0, "exposure": 0, "fixed": 0, "total": 0,
            "days_over": 0, "priced": p.regime != UNMAPPED_REGIME_LABEL,
        })
        s["planned"] += p.planned
        s["nonexec"] += p.nonexec
        if p.over_threshold:
            s["days_over"] += 1
        if s["priced"]:
            s["exposure"] += p.exposure or 0
            s["fixed"] += p.fixed_total or 0
            s["total"] += p.total or 0
    return sorted(agg.values(),
                  key=lambda s: (not s["priced"], -(s["total"] if s["priced"] else s["nonexec"])))


# ---------- 4. אי-דיוק — דגימת קו גדול ----------
def measure_inaccuracy(days: list[DayData], recs: list[LineDayRecord]) -> dict:
    """
    דוגם את הקו הגדול ביותר (לפי מתוכנן) ביום חול תקף, מושך rides_execution/list,
    ובודק אם actual_start_time שונה מ-planned_start_time. אם זהה בכל השורות -> לא מדיד.
    """
    valid_dates = {d.date for d in days if d.valid and d.is_workday}
    cand = [r for r in recs if r.date in valid_dates and r.planned > 0]
    if not cand:
        return {"status": "אין נתונים", "detail": "לא נמצאו ימי חול תקפים לדגימה"}
    big = max(cand, key=lambda r: r.planned)

    r = SESSION.get(f"{API}/rides_execution/list", params={
        "date_from": big.date.isoformat(), "date_to": big.date.isoformat(),
        "operator_ref": str(big.op), "line_ref": str(big.line),
        "order_by": "planned_start_time asc", "limit": 5000,
    }, timeout=120)
    r.raise_for_status()
    rows = r.json()
    n = len(rows)
    diffs = sum(1 for x in rows
                if x.get("actual_start_time") and x.get("planned_start_time")
                and x["actual_start_time"] != x["planned_start_time"])
    sample = f"{OPERATORS[big.op]} קו {big.line} ({big.date.isoformat()}, {n} נסיעות)"
    if n == 0:
        return {"status": "אין נתונים", "detail": f"אין רשומות rides_execution ל-{sample}",
                "sample": sample, "n": n, "diffs": 0}
    if diffs == 0:
        return {"status": "אי-דיוק לא מדיד",
                "detail": ("actual_start_time == planned_start_time בכל השורות "
                           "(שדה מנוון, מצב 6/2026) — לא ניתן למדוד איחור/הקדמה"),
                "sample": sample, "n": n, "diffs": 0}
    return {"status": "מדיד חלקית",
            "detail": f"{diffs}/{n} נסיעות עם זמן יציאה בפועל שונה מהמתוכנן",
            "sample": sample, "n": n, "diffs": diffs}


# ============================================================================
#                       5. ייצור xlsx (openpyxl)
# ============================================================================
def build_workbook(week: tuple[dt.date, dt.date], days: list[DayData],
                   recs: list[LineDayRecord], op_agg: dict,
                   penalties: list[PenaltyRow], inaccuracy: dict, path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    sun, sat = week
    ARIAL = "Arial"
    HDR_FILL = PatternFill("solid", fgColor="1F4E78")
    HDR_FONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name=ARIAL, bold=True, size=14, color="1F4E78")
    BASE_FONT = Font(name=ARIAL, size=10)
    WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
    BAD_FILL = PatternFill("solid", fgColor="F8CBAD")
    OK_FILL = PatternFill("solid", fgColor="E2EFDA")
    thin = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    PCT = "0.0%"; NUM = "#,##0"; KM = "#,##0"; SHK = '#,##0" ₪"'

    wb = Workbook()
    wb.remove(wb.active)

    def new_sheet(title: str) -> "object":
        ws = wb.create_sheet(title)
        ws.sheet_view.rightToLeft = True   # RTL לעברית
        return ws

    def header_row(ws, headers, row=1):
        for j, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=j, value=h)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.alignment = CENTER; c.border = BORDER
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    def style_body(ws, first_data_row, ncols):
        for row in ws.iter_rows(min_row=first_data_row, max_col=ncols):
            for c in row:
                if c.font is None or c.font.name != ARIAL:
                    c.font = BASE_FONT
                c.border = BORDER

    def autosize(ws, widths):
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w

    valid_workdays = [d for d in days if d.valid and d.is_workday]
    week_label = f"{sun.isoformat()} – {sat.isoformat()}"

    # ---- גיליון 1: סיכום מנהלים ----
    ws = new_sheet("סיכום מנהלים")
    ws.merge_cells("A1:H1")
    t = ws.cell(1, 1, f"דוח אי-ביצוע שבועי — אי-יציאה (2.1.1) | שבוע {week_label}")
    t.font = TITLE_FONT; t.alignment = CENTER
    ws.merge_cells("A2:H2")
    ws.cell(2, 1, "המדד מודד אי-יציאה בלבד = חסם תחתון לשיעור אי-הביצוע הרשמי (נספח כ\"ו)").font = \
        Font(name=ARIAL, italic=True, size=9, color="808080")
    headers = ["מפעיל", "מתוכנן (נסיעות)", "בוצע (נסיעות)", "% ביצוע", "% אי-ביצוע",
               "אי-ביצוע ק\"מ", "% אי-ביצוע ק\"מ", "ימי-חול תקפים"]
    header_row(ws, headers, row=4)
    r0 = 5
    for i, op in enumerate(OPERATORS):
        a = op_agg[op]; rr = r0 + i
        ws.cell(rr, 1, OPERATORS[op])
        ws.cell(rr, 2, a.get("planned", 0)).number_format = NUM
        ws.cell(rr, 3, a.get("executed", 0)).number_format = NUM
        # נוסחאות חיות
        ws.cell(rr, 4, f"=IF(B{rr}=0,0,C{rr}/B{rr})").number_format = PCT
        ws.cell(rr, 5, f"=IF(B{rr}=0,0,1-C{rr}/B{rr})").number_format = PCT
        ws.cell(rr, 6, round(a.get("nonexec_km", 0))).number_format = KM
        pkm = a.get("planned_km", 0)
        ws.cell(rr, 7, f"=IF({pkm}=0,0,F{rr}/{pkm})").number_format = PCT
        ws.cell(rr, 8, a.get("valid_days", 0)).alignment = CENTER
        # צביעת חריגה מסף שירות
        cell = ws.cell(rr, 5)
        nonexec = a.get("nonexec", 0); planned = a.get("planned", 0)
        rate = (nonexec / planned) if planned else 0
        cell.fill = BAD_FILL if rate > THRESH_FUNDAMENTAL else (WARN_FILL if rate > THRESH_SERVICE else OK_FILL)
    last = r0 + len(OPERATORS) - 1
    tot = last + 1
    ws.cell(tot, 1, "סך הכל").font = Font(name=ARIAL, bold=True)
    ws.cell(tot, 2, f"=SUM(B{r0}:B{last})").number_format = NUM
    ws.cell(tot, 3, f"=SUM(C{r0}:C{last})").number_format = NUM
    ws.cell(tot, 4, f"=IF(B{tot}=0,0,C{tot}/B{tot})").number_format = PCT
    ws.cell(tot, 5, f"=IF(B{tot}=0,0,1-C{tot}/B{tot})").number_format = PCT
    ws.cell(tot, 6, f"=SUM(F{r0}:F{last})").number_format = KM
    for j in range(1, 9):
        ws.cell(tot, j).fill = HDR_FILL; ws.cell(tot, j).font = Font(name=ARIAL, bold=True, color="FFFFFF")
    style_body(ws, 5, 8); autosize(ws, [16, 16, 16, 10, 12, 14, 14, 14])
    # גרף % אי-ביצוע פר-מפעיל
    chart = BarChart(); chart.type = "col"; chart.title = "% אי-ביצוע לפי מפעיל"
    chart.y_axis.numFmt = "0.0%"; chart.height = 8; chart.width = 18
    data = Reference(ws, min_col=5, min_row=4, max_row=last)
    cats = Reference(ws, min_col=1, min_row=r0, max_row=last)
    chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
    ws.add_chart(chart, f"J4")

    # ---- גיליון 2: אי-ביצוע יומי לפי מפעיל ----
    ws = new_sheet("יומי לפי מפעיל")
    ws.merge_cells("A1:I1"); ws.cell(1, 1, "אי-ביצוע יומי לפי מפעיל (% — נסיעות)").font = TITLE_FONT
    day_cols = [d for d in days]
    headers = ["מפעיל"] + [f"{HEB_DOW[d.dow]} {d.date.strftime('%d/%m')}" for d in day_cols]
    header_row(ws, headers, row=3)
    # מתוכנן/בוצע פר (op, יום)
    pe: dict[tuple[int, dt.date], list[int]] = collections.defaultdict(lambda: [0, 0])
    for r in recs:
        pe[(r.op, r.date)][0] += r.planned; pe[(r.op, r.date)][1] += r.executed
    for i, op in enumerate(OPERATORS):
        rr = 4 + i; ws.cell(rr, 1, OPERATORS[op]).font = Font(name=ARIAL, bold=True)
        for j, d in enumerate(day_cols, start=2):
            p, e = pe[(op, d.date)]
            cell = ws.cell(rr, j)
            if not d.percentages:
                cell.value = "—"; cell.alignment = CENTER
            elif p == 0:
                cell.value = "—"; cell.alignment = CENTER
            else:
                rate = 1 - e / p
                cell.value = rate; cell.number_format = PCT
                if not d.valid or op in d.strike_ops:
                    cell.fill = WARN_FILL  # יום/מפעיל מוחרג — מוצג אך מסומן
                elif rate > THRESH_FUNDAMENTAL:
                    cell.fill = BAD_FILL
    style_body(ws, 4, len(headers)); autosize(ws, [16] + [11] * len(day_cols))
    note_r = 4 + len(OPERATORS) + 1
    ws.cell(note_r, 1, "כתום = יום/מפעיל מוחרג מהממוצע;  אדום = מעל 2.5% (הפרה יסודית);  — = יום לא-תקף/ללא תכנון").font = \
        Font(name=ARIAL, italic=True, size=9, color="808080")

    # ---- גיליון 3: אי-ביצוע לפי אשכול ----
    ws = new_sheet("לפי אשכול")
    ws.merge_cells("A1:F1"); ws.cell(1, 1, "אי-ביצוע לפי אשכול (ימי-חול תקפים)").font = TITLE_FONT
    header_row(ws, ["אשכול", "מפעיל ראשי", "מתוכנן", "בוצע", "% ביצוע", "% אי-ביצוע"], row=3)
    cl: dict[str, collections.Counter] = collections.defaultdict(collections.Counter)
    cl_op: dict[str, collections.Counter] = collections.defaultdict(collections.Counter)
    valid_dates = {d.date for d in valid_workdays}
    strike = {(d.date, op) for d in days for op in d.strike_ops}
    for r in recs:
        if r.date not in valid_dates or (r.date, r.op) in strike:
            continue
        cl[r.cluster]["planned"] += r.planned; cl[r.cluster]["executed"] += r.executed
        cl_op[r.cluster][r.op_name] += r.planned
    rr = 4
    for cluster in sorted(cl, key=lambda c: -cl[c]["planned"]):
        b = cl[cluster]; main_op = cl_op[cluster].most_common(1)[0][0] if cl_op[cluster] else ""
        ws.cell(rr, 1, cluster); ws.cell(rr, 2, main_op)
        ws.cell(rr, 3, b["planned"]).number_format = NUM
        ws.cell(rr, 4, b["executed"]).number_format = NUM
        ws.cell(rr, 5, f"=IF(C{rr}=0,0,D{rr}/C{rr})").number_format = PCT
        ws.cell(rr, 6, f"=IF(C{rr}=0,0,1-D{rr}/C{rr})").number_format = PCT
        rate = (1 - b["executed"] / b["planned"]) if b["planned"] else 0
        ws.cell(rr, 6).fill = BAD_FILL if rate > THRESH_FUNDAMENTAL else (WARN_FILL if rate > THRESH_SERVICE else OK_FILL)
        rr += 1
    style_body(ws, 4, 6); autosize(ws, [22, 14, 14, 14, 10, 12])

    # ---- גיליון 4: חשיפת קנסות — פר-משטר-מכרז ----
    ws = new_sheet("חשיפת קנסות")
    ws.merge_cells("A1:M1")
    ws.cell(1, 1, "חשיפת קנסות לפי מפעיל×אשכול×יום — פר-משטר-מכרז "
                  "(נספח כ\"ו; סכומים נומינליים, ללא הצמדת מדד 2/2012)").font = TITLE_FONT
    headers = ["תאריך", "יום", "מפעיל", "אשכול", "משטר", "מתוכנן", "אי-בוצע",
               "% אי-ביצוע", "תעריף מדורג ₪/הפרה", "חשיפה מדורגת ₪",
               "רכיבי קו/נסיעה ₪", "סה\"כ ₪", "פירוט רכיבים / לא נמצא במקור"]
    header_row(ws, headers, row=3)
    ws.cell(3, 5).comment = Comment(
        "מנוע-קנסות פר-משטר-מכרז (TENDER_REGIMES): בקעת אונו אלעד=מכרז 5/2021; "
        "שרון=מכרז 04/2021; שרון חולון מרחבי=מכרז 24/2015; הנגב=מכרז 07/2014. "
        "המשטרים חלים על מטרופולין בלבד — שורת מפעיל אחר וכן אשכול אחר "
        "מסומנים 'לא-ממופה' ואינם מתומחרים (לא 0 בשקט). "
        "הנתיב האשכולי הישן של 5,000 ₪ נותר מנוטרל.", "בקרה")
    NA = "—"
    rr = 4
    for p in penalties:
        unmapped = p.regime == UNMAPPED_REGIME_LABEL
        ws.cell(rr, 1, p.date.isoformat()); ws.cell(rr, 2, p.dow).alignment = CENTER
        ws.cell(rr, 3, p.op_name); ws.cell(rr, 4, p.cluster)
        ws.cell(rr, 5, p.regime)
        ws.cell(rr, 6, p.planned).number_format = NUM
        ws.cell(rr, 7, p.nonexec).number_format = NUM
        ws.cell(rr, 8, f"=IF(F{rr}=0,0,G{rr}/F{rr})").number_format = PCT
        if p.tariff is None:
            ws.cell(rr, 9, NA).alignment = CENTER
            ws.cell(rr, 10, NA).alignment = CENTER
        else:
            ws.cell(rr, 9, p.tariff).number_format = SHK
            ws.cell(rr, 10, f"=G{rr}*I{rr}").number_format = SHK
        if unmapped:
            ws.cell(rr, 11, NA).alignment = CENTER
            ws.cell(rr, 12, "לא-ממופה").alignment = CENTER
        else:
            ws.cell(rr, 11, p.fixed_total).number_format = SHK   # K כערך
            # סה"כ (L) = נוסחה על J+K; במשטר ללא מדרג J מכיל "—" ולכן =K בלבד
            total_formula = f"=J{rr}+K{rr}" if p.tariff is not None else f"=K{rr}"
            ws.cell(rr, 12, total_formula).number_format = SHK
        detail = [f"{label}: {amount:,} ₪" for label, amount in (p.components or [])]
        if p.not_found:
            detail.append(p.not_found)
        ws.cell(rr, 13, " | ".join(detail) or NA)
        if p.over_threshold:
            for j in range(1, 14):
                ws.cell(rr, j).fill = BAD_FILL
        elif unmapped:
            for j in range(1, 14):
                ws.cell(rr, j).fill = WARN_FILL
        rr += 1
    if rr > 4:
        last = rr - 1
        ws.cell(rr, 4, "סך חשיפה (משטרים ממופים בלבד)").font = Font(name=ARIAL, bold=True)
        ws.cell(rr, 10, f"=SUM(J4:J{last})").number_format = SHK
        ws.cell(rr, 11, f"=SUM(K4:K{last})").number_format = SHK
        ws.cell(rr, 12, f"=SUM(L4:L{last})").number_format = SHK
        for j in range(1, 14):
            ws.cell(rr, j).fill = HDR_FILL; ws.cell(rr, j).font = Font(name=ARIAL, bold=True, color="FFFFFF")
        rr += 1
    style_body(ws, 4, 13)
    autosize(ws, [11, 5, 14, 18, 26, 10, 10, 10, 13, 14, 14, 13, 60])

    # סיכום פר-אשכול (rule: אשכול לא-ממופה מסומן, לא מתומחר 0 בשקט)
    rr += 2
    ws.cell(rr, 1, "סיכום חשיפה פר-אשכול").font = TITLE_FONT
    rr += 1
    sum_headers = ["אשכול", "משטר", "מתוכנן", "אי-בוצע", "חשיפה מדורגת ₪",
                   "רכיבי קו/נסיעה ₪", "סה\"כ ₪", "שורות מפעיל-יום מעל 4.5%", "הערה"]
    header_row(ws, sum_headers, row=rr)
    first_sum = rr + 1
    rr += 1
    for s in summarize_penalties_by_cluster(penalties):
        ws.cell(rr, 1, s["cluster"]); ws.cell(rr, 2, s["regime"])
        ws.cell(rr, 3, s["planned"]).number_format = NUM
        ws.cell(rr, 4, s["nonexec"]).number_format = NUM
        if s["priced"]:
            ws.cell(rr, 5, s["exposure"]).number_format = SHK
            ws.cell(rr, 6, s["fixed"]).number_format = SHK
            ws.cell(rr, 7, s["total"]).number_format = SHK
            ws.cell(rr, 9, NA)
        else:
            ws.cell(rr, 5, NA).alignment = CENTER
            ws.cell(rr, 6, NA).alignment = CENTER
            ws.cell(rr, 7, NA).alignment = CENTER
            ws.cell(rr, 9, "לא-ממופה — האשכול אינו משויך למשטר-מכרז, החשיפה לא תומחרה")
            for j in range(1, 10):
                ws.cell(rr, j).fill = WARN_FILL
        ws.cell(rr, 8, s["days_over"]).alignment = CENTER
        rr += 1
    style_body(ws, first_sum, 9)
    ws.freeze_panes = "A4"   # header_row של הסיכום מזיז את ההקפאה — החזרה לכותרת הפירוט

    # ---- גיליון 5: השוואת P95 ----
    ws = new_sheet("השוואת P95")
    ws.merge_cells("A1:H1")
    ws.cell(1, 1, "אחוז שירות שלא הופעל מול בסיס P95 (יום-חול ממוצע)").font = TITLE_FONT
    headers = ["מפעיל", "בסיס P95 נסיעות", "ממוצע יומי בוצע", "% שלא הופעל (נסיעות)",
               "בסיס P95 ק\"מ", "ממוצע יומי בוצע ק\"מ", "% שלא הופעל (ק\"מ)"]
    header_row(ws, headers, row=3)
    rr = 4
    for op in OPERATORS:
        a = op_agg[op]; vd = max(a.get("valid_days", 0), 1)
        p95_rides, p95_km = P95[op]
        avg_rides = a.get("executed", 0) / vd
        avg_km = a.get("executed_km", 0) / vd
        ws.cell(rr, 1, OPERATORS[op])
        ws.cell(rr, 2, p95_rides).number_format = NUM
        ws.cell(rr, 3, round(avg_rides)).number_format = NUM
        ws.cell(rr, 4, f"=IF(B{rr}=0,0,1-C{rr}/B{rr})").number_format = PCT
        ws.cell(rr, 5, p95_km).number_format = KM
        ws.cell(rr, 6, round(avg_km)).number_format = KM
        ws.cell(rr, 7, f"=IF(E{rr}=0,0,1-F{rr}/E{rr})").number_format = PCT
        rr += 1
    style_body(ws, 4, 7); autosize(ws, [16, 16, 16, 16, 16, 18, 16])
    ws.cell(rr + 1, 1, "% שלא הופעל = 1 − בוצע/P95.  P95 = אחוזון 95 של יום-חול (חושב 10.6.2026).").font = \
        Font(name=ARIAL, italic=True, size=9, color="808080")

    # ---- גיליון 6: אי-דיוק ----
    ws = new_sheet("אי-דיוק")
    ws.merge_cells("A1:D1"); ws.cell(1, 1, "אי-דיוק (סעיף 2.5) — דגימה").font = TITLE_FONT
    rows6 = [
        ("סטטוס", inaccuracy.get("status", "")),
        ("קו שנדגם", inaccuracy.get("sample", "")),
        ("מספר נסיעות בדגימה", inaccuracy.get("n", "")),
        ("נסיעות עם זמן יציאה בפועל שונה", inaccuracy.get("diffs", "")),
        ("פירוט", inaccuracy.get("detail", "")),
    ]
    for i, (k, v) in enumerate(rows6, start=3):
        ws.cell(i, 1, k).font = Font(name=ARIAL, bold=True)
        ws.cell(i, 2, v).font = BASE_FONT
    autosize(ws, [26, 80])

    # ---- גיליון 7: ימים חריגים / החרגות ----
    ws = new_sheet("ימים חריגים")
    ws.merge_cells("A1:E1"); ws.cell(1, 1, "סיווג ימים בשבוע הדיווח").font = TITLE_FONT
    header_row(ws, ["תאריך", "יום", "סיווג", "נכלל בממוצע?", "סיבה / הערה"], row=3)
    rr = 4
    for d in days:
        ws.cell(rr, 1, d.date.isoformat()); ws.cell(rr, 2, HEB_DOW[d.dow]).alignment = CENTER
        ws.cell(rr, 3, d.classification)
        ws.cell(rr, 4, "כן" if (d.valid and d.is_workday and not d.strike_ops) else "חלקי" if d.strike_ops and d.valid else "לא").alignment = CENTER
        ws.cell(rr, 5, d.reason or "—")
        if not d.valid:
            for j in range(1, 6):
                ws.cell(rr, j).fill = WARN_FILL
        rr += 1
    style_body(ws, 4, 5); autosize(ws, [12, 5, 22, 14, 60])

    # ---- גיליון 8: נתוני גלם — מתוכנן מול בוצע לפי קו ----
    ws = new_sheet("גלם — קו×יום")
    header_row(ws, ["תאריך", "יום", "מפעיל", "אשכול", "route_mkt", "line_ref",
                    "מתוכנן", "בוצע", "אי-בוצע", "% אי-ביצוע"], row=1)
    rr = 2
    for r in sorted(recs, key=lambda x: (x.date, x.op, -x.nonexec)):
        ws.cell(rr, 1, r.date.isoformat()); ws.cell(rr, 2, r.dow).alignment = CENTER
        ws.cell(rr, 3, r.op_name); ws.cell(rr, 4, r.cluster)
        ws.cell(rr, 5, r.mkt); ws.cell(rr, 6, r.line)
        ws.cell(rr, 7, r.planned).number_format = NUM
        ws.cell(rr, 8, r.executed).number_format = NUM
        ws.cell(rr, 9, r.nonexec).number_format = NUM
        ws.cell(rr, 10, f"=IF(G{rr}=0,0,I{rr}/G{rr})").number_format = PCT
        rr += 1
    autosize(ws, [11, 5, 14, 18, 10, 10, 9, 9, 9, 11])

    # ---- גיליון 9: ק"מ — מתוכנן מול בוצע ----
    ws = new_sheet("ק\"מ — קו×יום")
    header_row(ws, ["תאריך", "יום", "מפעיל", "line_ref", "ק\"מ/נסיעה",
                    "מתוכנן ק\"מ", "בוצע ק\"מ", "אי-בוצע ק\"מ"], row=1)
    rr = 2
    for r in sorted(recs, key=lambda x: (x.date, x.op, -x.nonexec_km)):
        if r.km_per_ride == 0:
            continue
        ws.cell(rr, 1, r.date.isoformat()); ws.cell(rr, 2, r.dow).alignment = CENTER
        ws.cell(rr, 3, r.op_name); ws.cell(rr, 4, r.line)
        ws.cell(rr, 5, r.km_per_ride).number_format = "0.0"
        ws.cell(rr, 6, round(r.planned_km, 1)).number_format = KM
        ws.cell(rr, 7, round(r.executed_km, 1)).number_format = KM
        ws.cell(rr, 8, round(r.nonexec_km, 1)).number_format = KM
        rr += 1
    autosize(ws, [11, 5, 14, 10, 11, 13, 13, 13])

    # ---- גיליון 10: מתודולוגיה והגדרות ----
    ws = new_sheet("מתודולוגיה")
    ws.merge_cells("A1:B1"); ws.cell(1, 1, "מתודולוגיה, הגדרות וסייגים").font = TITLE_FONT
    lines = [
        ("מקור הגדרות", "נספח כ\"ו (פיצויים מוסכמים) להסכם ההפעלה הסטנדרטי, משרד התחבורה."),
        ("מה נמדד", "אי-יציאה (2.1.1) בלבד: נסיעה מתוכננת (GTFS) ללא נסיעת SIRI תואמת."),
        ("חסם תחתון", "השיעור הרשמי גבוה יותר — כולל גם איחור חמור (2.1.2), הקדמה חמורה (2.1.3) וחפיפה (2.1.4)."),
        ("מתוכנן", "/gtfs_rides_agg/group_by לפי operator_ref,line_ref — total_planned_rides."),
        ("בוצע", "/siri_rides — דה-דופ לפי journey_ref (journey שמסתיים ב-\"-0\" נספר בנפרד); בוצע ≤ מתוכנן לכל קו."),
        ("הצלבת אשכול", "line_ref → route_mkt (/gtfs_routes) → OfficeLineId ב-ClusterToLine.zip (ללא אפסים מובילים)."),
        ("ק\"מ", "אורך ה-shape השכיח לכל route_id ב-GTFS הארצי (Haversine)."),
        ("ימי חול", "ראשון–חמישי בלבד. שישי/שבת — דפוס בסיס שונה, מוצגים בנפרד."),
        ("יום מופחת", f"מתוכנן < {int(REDUCED_PLAN_RATIO*100)}% מחציון יום-חול אצל רוב המפעילים — מוחרג מהממוצע."),
        ("חריג ביצוע", f"תכנון רגיל אך ביצוע מפעיל < {int(STRIKE_EXEC_RATIO*100)}% מהרגיל — המפעיל מוחרג ליום זה."),
        ("תקלת SIRI", "ביצוע נמוך חוצה-מפעילים + נפילת מצוק שעתית לאפס — סומן 'לא תקף', ללא אחוזים, הוחרג מהכול."),
        ("כפל תכנון GTFS", f"מתוכנן קו ≥ פי {DOUBLE_PLAN_RATIO} מחציון יום-חול — הוצב חציון (אחיד על כל ימי החול)."),
        ("ספי שירות", "אי-ביצוע > 2.1% חריגה משירות; > 2.5% הפרה יסודית; אי-דיוק > 4.5% חריגה."),
        ("קנסות — פר-משטר-מכרז",
         "בקעת אונו אלעד (מכרז 5/2021) ושרון (מכרז 04/2021): מדרג אשכולי 0/63/93/118/143/173 ₪ להפרה "
         "לפי שיעור אי-הביצוע היומי באשכול + קנס קו-יום לקו רגיל מעל 4.5% (אונו: 5,000 ₪ §2.2; "
         "שרון: 500 ₪ לתקופת-יום §2.1 — חסם תחתון של תקופה אחת מתוך עד 3, באין פיצול תקופות בנתונים) "
         "+ 5,000 ₪ לכל נסיעה שלא בוצעה בקו בתדירות נמוכה (§2.3/§2.2). "
         "שרון חולון מרחבי (24/2015) והנגב (07/2014): 2,000 ₪ לכל נסיעה שלא בוצעה — "
         "ללא כלל 4.5% וללא מדרג; לפני מקדם ההכפלה שנקב המפעיל (לא נמצא במקורות; מקדם ≥ 1 לא הוחל). "
         "המשטרים חלים על אשכולות מטרופולין בלבד; שורת מפעיל אחר ואשכול לא-ממופה אינם מתומחרים "
         "ומסומנים 'לא-ממופה'. הסכומים נומינליים, ללא הצמדת מדד 2/2012."),
        ("משטרים ישנים — אופי החשיפה",
         "במכרזי 24/2015 ו-07/2014 הטריגר החוזי לקנס הוא אירוע מתועד (דיווח פקח משרד התחבורה / "
         "מוסמך מטעם המפקח / תלונת ציבור מוצדקת), לא בקרה אלקטרונית על כלל הנסיעות. "
         "התמחור פר-נסיעת-SIRI שלא בוצעה הוא לכן תקרת-חשיפה תיאורטית, לא חסם תחתון."),
        ("קו בתדירות נמוכה",
         "קו שביום נתון 100% מנסיעותיו הן בתדירות 59 דק' ומעלה — כלומר המרווח-העוקב המינימלי "
         "בין כל זוג נסיעות סמוכות (מזמני היציאה המתוכננים שנצפו ב-SIRI) ≥ 59 דק'; "
         "קו עם נסיעה מתוכננת אחת ביום = תדירות נמוכה. כל נסיעותיו 'לא תדירות' — "
         "מנותבות לסעיף המחמיר (5,000 ₪/נסיעה) במקום הקנס הקבוע של קו רגיל. "
         "לא-ניתן-לקביעה (זמנים חסרים/חלקיים, או יציאות בחלון 00:00-03:59 שמעיד על זיהום "
         "חלון ה-UTC היומי בזנבות יום-השירות הקודם/הבא) — הקו מטופל כקו רגיל ומדווח."),
        ("רכיבים שלא תומחרו",
         "רכיב שסכומו לא נלכד במקורות (למשל נסיעה-אחרונה בשרון, מקדמי ההכפלה במכרזים הישנים) "
         "אינו מתומחר ומסומן בעמודת 'פירוט רכיבים / לא נמצא במקור'; רכיב שאינו מדיד מהנתונים "
         "(איחור/הקדמה, נסיעה אחרונה, תלונות ציבור) מתועד ב-TENDER_REGIMES ואינו מתומחר — "
         "לכן במשטרי 2021 (בקרה אלקטרונית) החשיפה המדווחת היא חסם תחתון; במשטרים הישנים "
         "(24/2015, 07/2014) היא תקרת-חשיפה תיאורטית — ר' 'משטרים ישנים — אופי החשיפה'."),
        ("אי-דיוק", inaccuracy.get("detail", "")),
    ]
    seen_regimes: set[str] = set()
    for regime in TENDER_REGIMES.values():
        if regime["name"] in seen_regimes:
            continue
        seen_regimes.add(regime["name"])
        items = list(regime.get("not_found", [])) + list(regime.get("not_computable", []))
        if items:
            lines.append((f"לא תומחר — {regime['name']}", " · ".join(items)))
    rr = 3
    for k, v in lines:
        ws.cell(rr, 1, k).font = Font(name=ARIAL, bold=True, color="1F4E78")
        c = ws.cell(rr, 2, v); c.font = BASE_FONT; c.alignment = Alignment(wrap_text=True, vertical="top")
        rr += 1
    autosize(ws, [20, 100])
    ws.column_dimensions["B"].width = 100

    wb.save(path)


# ============================================================================
#                                main
# ============================================================================
def run(week_ending: dt.date | None, output_dir: str) -> str:
    import os
    today = dt.date.today()
    if week_ending:
        sat = week_ending
        while sat.weekday() != 5:
            sat -= dt.timedelta(days=1)
        sun = sat - dt.timedelta(days=6)
    else:
        sun, sat = week_ending_saturday(today)
    week = (sun, sat)
    dates = [sun + dt.timedelta(days=i) for i in range(7)]
    print(f"[i] שבוע דיווח: {sun} .. {sat}")

    # מקורות עזר (פעם אחת)
    print("[i] טוען לוח החרגות, אשכולות (ClusterToLine), ואורכי GTFS…")
    calendar_entries = load_calendar()
    clusters = load_clusters()
    # איחוד שבת + יום-חול: שבת לבדה מחמיצה כמחצית מהקווים (ראו fetch_line_to_mkt).
    weekday_ref = next((x for x in reversed(dates) if x.weekday() in WEEKDAY_WORK), sat)
    line_mkt = fetch_line_to_mkt([sat, weekday_ref])
    route_len = load_gtfs_lengths()
    print(f"[i] אשכולות: {len(set(clusters.values()))}, קווי-mkt ממופים: {len(line_mkt)}, "
          f"קווים עם אורך: {len(route_len)}")

    # משיכת ימים
    print("[i] מושך מתוכנן + בוצע (SIRI) לכל יום…")
    days: list[DayData] = []
    for d in dates:
        day = fetch_day(d, line_mkt)
        days.append(day)
        print(f"    {HEB_DOW[d.weekday()]} {d}: "
              f"מתוכנן={sum(day.planned_raw.values())}, "
              f"בוצע={sum(sum(v.values()) for v in day.executed.values())}")

    # סיווג ימים חריגים
    print("[i] מסווג ימים חריגים…")
    classify_days(days, calendar_entries)
    for d in days:
        flag = "" if (d.valid and d.is_workday) else f"  ← {d.classification}"
        print(f"    {HEB_DOW[d.dow]} {d.date}: {d.classification}{flag}")

    # רשומות, אגרגציה, קנסות, אי-דיוק
    recs = build_records(days, line_mkt, clusters, route_len)
    op_agg = aggregate_operator(days, recs)
    penalties = compute_penalties(days, recs)
    inaccuracy = measure_inaccuracy(days, recs)
    print(f"[i] רשומות קו×יום: {len(recs)} | שורות קנס: {len(penalties)} | "
          f"אי-דיוק: {inaccuracy.get('status')}")

    # ייצור xlsx
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, f"דוח_שבועי_אי_ביצוע_{sat.isoformat()}.xlsx")
    print("[i] בונה xlsx…")
    build_workbook(week, days, recs, op_agg, penalties, inaccuracy, path)
    print(f"[✓] נשמר: {path}")
    return path


def main() -> None:
    ap = argparse.ArgumentParser(description="דוח אי-ביצוע שבועי לתח\"צ בישראל (אי-יציאה 2.1.1)")
    ap.add_argument("--week-ending", type=lambda s: dt.date.fromisoformat(s),
                    default=None, help="תאריך סיום שבוע (שבת). ברירת מחדל: השבוע שהסתיים אתמול.")
    ap.add_argument("--output-dir", default="outputs", help="תיקיית פלט ל-xlsx.")
    args = ap.parse_args()
    path = run(args.week_ending, args.output_dir)
    print(path)


if __name__ == "__main__":
    main()
