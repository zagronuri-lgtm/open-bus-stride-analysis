"""חוברת נתוני אמת לסניף מטרופולין 500 — Power BI קודם, Stride להשוואה בלבד.

אין שליפת BI חיה בסביבת cloud (אין browser). העוגנים השמורים מכיול 08/07
הם מקור האמת הזמין; גיליון «הדבק BI חי» מיועד לרענון כשיש גישה לדוח.

הרצה:
  python -m src.branch_500_ground_truth \\
    --investigation-json outputs/חקירת_סניף_500_2026-07-04.json \\
    --output-dir outputs
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

POWERBI_URL = (
    "https://app.powerbi.com/groups/me/reports/"
    "cae01515-0e87-42dc-b643-244528bc2bca/bf6381d728e43ccdcd08"
    "?experience=power-bi"
)

# נתוני אמת שמורים — לא Stride
BI_TRUTH = {
    "aggregate_branch_500": {
        "label": "סניף 500 — אגרגט (כיול 08/07)",
        "pct": 0.0253,
        "filters": "Branch=500 (אגרגט מצורף לכיול)",
        "note": "זה המספר התפעולי/רשמי שהמשתמש מדווח כ־2–3%",
        "source": "METROPOLIN_BRANCHES_AGG / create_calibration_workbook.py",
    },
    "day_2026_07_05": {
        "label": "סניף 500 — יום 05/07/2026",
        "pct": 0.0531,
        "filters": "Branch=500, יום בודד",
        "note": "באותו יום Stride=BI=5.31% — אין פער הגדרה בחלון זהה",
        "source": "עוגן מצורף לכיול 08/07",
    },
    "metropoline_july_mtd_to_07": {
        "label": "מטרופולין כולה — יולי עד כספת 07/07",
        "rides": 69_644,
        "rides_for_calc": 68_717,
        "non_execution": 1_234,
        "pct": 1_234 / 68_717,
        "filters": "Month=2026-07 Branch=All Cluster=All",
        "note": "נשלף חי 09/07; לא סניף 500 בלבד",
        "source": "Power BI live 2026-07-09",
    },
}


def _styles():
    return {
        "title": Font(name="Arial", bold=True, size=14, color="1F4E78"),
        "hdr_fill": PatternFill("solid", fgColor="1F4E78"),
        "hdr_font": Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "base": Font(name="Arial", size=10),
        "truth": PatternFill("solid", fgColor="C6EFCE"),
        "compare": PatternFill("solid", fgColor="FFF2CC"),
        "warn": PatternFill("solid", fgColor="FCE4D6"),
    }


def _header(ws, headers, row=1, styles=None):
    s = styles or _styles()
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.fill = s["hdr_fill"]
        c.font = s["hdr_font"]
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def load_mapping_500() -> list[dict]:
    path = Path("data/reference/metropoline_line_branch_map.csv")
    rows = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            if (row.get("branch") or "").strip() == "500":
                rows.append(row)
    return rows


def build_workbook(investigation: dict | None, output_path: Path) -> Path:
    s = _styles()
    wb = Workbook()
    wb.remove(wb.active)

    # ---- 1. נתוני אמת (BI) ----
    ws = wb.create_sheet("נתוני אמת — BI", 0)
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "סניף 500 — נתוני אמת (Power BI)"
    ws["A1"].font = s["title"]
    ws["A2"] = (
        "מקור האמת הוא Power BI / בקרה רשמית. Stride אינו מחליף את המספרים האלה — "
        "הוא חסם תחתון (אי-יציאה 2.1.1) להשוואה בלבד."
    )
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
    _header(ws, ["פריט", "ערך", "מסננים / הערה", "מקור"], row=4, styles=s)

    rr = 5
    for key, a in BI_TRUTH.items():
        ws.cell(rr, 1, a["label"]).font = Font(name="Arial", bold=True)
        if "rides" in a:
            ws.cell(rr, 2, f"נסיעות={a['rides']:,} | לחישוב={a['rides_for_calc']:,} | "
                           f"אי-ביצוע={a['non_execution']:,} | %{a['pct']:.2%}")
        else:
            cell = ws.cell(rr, 2, a["pct"])
            cell.number_format = "0.00%"
            cell.fill = s["truth"]
        ws.cell(rr, 3, f"{a.get('filters', '')} — {a.get('note', '')}")
        ws.cell(rr, 4, a.get("source", ""))
        for j in range(1, 5):
            if ws.cell(rr, j).fill.fgColor is None or ws.cell(rr, j).fill.fgColor.rgb in (
                "00000000", None
            ):
                if j == 2 and "rides" not in a:
                    pass
                elif key.startswith("aggregate") or key.startswith("day_"):
                    if j != 2:
                        ws.cell(rr, j).fill = s["truth"]
        rr += 1

    rr += 1
    ws.cell(rr, 1, "קישור Power BI").font = Font(name="Arial", bold=True)
    ws.cell(rr, 2, POWERBI_URL)
    rr += 1
    ws.cell(rr, 1, "איך לרענן אמת חיה")
    ws.cell(
        rr, 2,
        "Reset → Month=2026-07 → Branch=500 → Cluster=All → העתק נסיעות / נסיעות לחישוב / "
        "אי ביצוע / תאריך כספת לגיליון «הדבק BI חי»",
    )
    _autosize(ws, [42, 55, 55, 40])

    # ---- 2. הדבק BI חי ----
    ws = wb.create_sheet("הדבק BI חי")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "הדבק כאן KPI חי מ-Power BI לסניף 500"
    ws["A1"].font = s["title"]
    _header(ws, ["שדה", "ערך", "הערה"], row=3, styles=s)
    fields = [
        ("Month", "2026-07", ""),
        ("Branch", "500", "חובה"),
        ("Cluster", "All", ""),
        ("תאריך כספת אחרון", "", ""),
        ("נסיעות", "", "B8"),
        ("נסיעות לחישוב", "", "B9 — מכנה"),
        ("החרגות", "", "אופציונלי"),
        ("אי ביצוע", "", "B11 — מונה"),
        ("איחור", "", "אופציונלי"),
        ("הקדמה", "", "אופציונלי"),
        ("% רשמי BI", None, "אי ביצוע / נסיעות לחישוב"),
        ("מקדם מכנה", None, "נסיעות לחישוב / נסיעות"),
    ]
    for i, (field, default, note) in enumerate(fields, start=4):
        ws.cell(i, 1, field).font = Font(name="Arial", bold=True)
        ws.cell(i, 2, default)
        ws.cell(i, 2).fill = s["warn"]
        ws.cell(i, 3, note).font = Font(name="Arial", italic=True, size=9, color="808080")
    ws["B14"] = '=IF(OR(B9="",B9=0),"",B11/B9)'
    ws["B14"].number_format = "0.00%"
    ws["B15"] = '=IF(OR(B8="",B8=0),"",B9/B8)'
    ws["B15"].number_format = "0.000"
    ws["A17"] = "אחרי מילוי — זה נתון האמת העדכני לסניף 500. אל תחליף בערכי Stride."
    ws["A17"].font = Font(name="Arial", bold=True, color="C65911")
    _autosize(ws, [22, 18, 40])

    # ---- 3. השוואת Stride (משני) ----
    ws = wb.create_sheet("השוואת Stride (משני)")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "Stride — חסם תחתון בלבד (לא אמת רשמית)"
    ws["A1"].font = s["title"]
    _header(ws, ["חלון Stride", "מתוכנן", "אי-בוצע", "% Stride", "% אמת BI", "פער נק׳ %"], row=3, styles=s)

    bi_agg = BI_TRUTH["aggregate_branch_500"]["pct"]
    windows = {}
    if investigation:
        windows = investigation.get("summary", {}).get("windows", {})
    labels = [
        ("valid_weekdays_sun_tue", "א׳–ג׳ תקפים (דוח שבועי)"),
        ("weekdays_sun_thu_incl_segment", "א׳–ה׳ כולל חופש גדול"),
        ("full_week", "כל השבוע"),
        ("july_only_wed_thu", "רק 01–02/07"),
    ]
    rr = 4
    for key, label in labels:
        w = windows.get(key, {})
        planned = w.get("planned", 0)
        missing = w.get("missing", 0)
        pct = w.get("pct", 0.0)
        ws.cell(rr, 1, label)
        ws.cell(rr, 2, planned).number_format = "#,##0"
        ws.cell(rr, 3, missing).number_format = "#,##0"
        c = ws.cell(rr, 4, pct)
        c.number_format = "0.00%"
        c.fill = s["compare"]
        c2 = ws.cell(rr, 5, bi_agg)
        c2.number_format = "0.00%"
        c2.fill = s["truth"]
        ws.cell(rr, 6, bi_agg - pct).number_format = "0.00%"
        rr += 1
    ws.cell(rr + 1, 1, "מסקנה").font = Font(name="Arial", bold=True, color="1F4E78")
    ws.cell(
        rr + 1, 2,
        "האמת ≈2.53%+. Stride נמוך יותר בגלל חלון (החרגת חופש גדול) + הגדרה 2.1.1 בלבד. "
        "המיפוי (20 מק״ט) יציב — לא באג שיוך.",
    )
    _autosize(ws, [32, 12, 12, 12, 12, 12])

    # ---- 4. מיפוי ----
    ws = wb.create_sheet("מיפוי 20 מק״ט")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "קווי סניף 500 במיפוי המקומי — נבדק: יציב, לא באג"
    ws["A1"].font = s["title"]
    _header(ws, ["route_mkt", "קו", "אשכול", "מוצא/יעד", "ייחודיות"], row=3, styles=s)
    for i, row in enumerate(load_mapping_500(), start=4):
        ws.cell(i, 1, row.get("route_mkt"))
        ws.cell(i, 2, row.get("route_short_name"))
        ws.cell(i, 3, row.get("cluster"))
        ws.cell(i, 4, row.get("od"))
        ws.cell(i, 5, row.get("line_uniqueness"))
    _autosize(ws, [12, 8, 22, 30, 18])

    # ---- 5. פירוט יומי/מק״ט מ-Stride ----
    if investigation:
        by_day = investigation.get("summary", {}).get("by_day", {})
        ws = wb.create_sheet("Stride פירוט יומי")
        ws.sheet_view.rightToLeft = True
        ws["A1"] = "פירוט יומי Stride לסניף 500 (להשוואה בלבד)"
        ws["A1"].font = s["title"]
        _header(ws, ["תאריך", "תקף?", "מתוכנן", "בוצע", "אי-בוצע", "%", "סיווג"], row=3, styles=s)
        rr = 4
        for day, slot in by_day.items():
            ws.cell(rr, 1, day)
            ws.cell(rr, 2, "כן" if slot.get("valid") else "לא")
            ws.cell(rr, 3, slot.get("planned", 0))
            ws.cell(rr, 4, slot.get("executed", 0))
            ws.cell(rr, 5, slot.get("missing", 0))
            c = ws.cell(rr, 6, slot.get("pct", 0))
            c.number_format = "0.00%"
            if not slot.get("valid"):
                c.fill = s["warn"]
            ws.cell(rr, 7, slot.get("classification", ""))
            rr += 1
        _autosize(ws, [12, 8, 12, 12, 12, 10, 28])

        by_mkt = investigation.get("summary", {}).get("by_mkt_valid", {})
        ws = wb.create_sheet("Stride מק״ט תקפים")
        ws.sheet_view.rightToLeft = True
        ws["A1"] = "מק״ט סניף 500 בימי חול תקפים (Stride)"
        ws["A1"].font = s["title"]
        _header(ws, ["route_mkt", "line_ref", "מתוכנן", "בוצע", "אי-בוצע", "%"], row=3, styles=s)
        rr = 4
        for mkt, slot in by_mkt.items():
            ws.cell(rr, 1, mkt)
            ws.cell(rr, 2, slot.get("line_ref", ""))
            ws.cell(rr, 3, slot.get("planned", 0))
            ws.cell(rr, 4, slot.get("executed", 0))
            ws.cell(rr, 5, slot.get("missing", 0))
            c = ws.cell(rr, 6, slot.get("pct", 0))
            c.number_format = "0.00%"
            if slot.get("pct", 0) >= 0.5:
                c.fill = s["warn"]
            rr += 1
        _autosize(ws, [12, 10, 12, 12, 12, 10])

    # ---- 6. מסקנות ----
    ws = wb.create_sheet("מסקנות חקירה")
    ws.sheet_view.rightToLeft = True
    ws["A1"] = "מסקנות — חקירת פער סניף 500"
    ws["A1"].font = s["title"]
    bullets = [
        "1. נתון האמת לסניף 500: ≈2.53% (אגרגט BI) / עד ~5.31% ביום חריג (05/07).",
        "2. דוח Stride השבועי (0.41%) אינו סותר את האמת — הוא חסם תחתון בחלון צר (א׳–ג׳ תקפים).",
        "3. גם אם כוללים חופש גדול, Stride נשאר ~0.9–1.1% — עדיין מתחת ל-BI בגלל הגדרה 2.1.1 בלבד.",
        "4. מיפוי 20 מק״ט יציב; אין עדות לבאג שיוך סניף.",
        "5. קווי לילה 11230/11231 תורמים חסרים יחסיים גבוהים בנפח קטן — לא מסבירים לבד את פער ה-2%.",
        "6. להשוואה תפעולית: להשתמש ב-Power BI Branch=500 (או גיליון «הדבק BI חי»), לא בסיכום מנהלים של Stride.",
        "7. שליפת BI חיה Branch=500 לא בוצעה בסביבת cloud זו — לרענן בגיליון ההדבקה כשיש דפדפן.",
    ]
    for i, t in enumerate(bullets, start=3):
        ws.cell(i, 1, t).font = s["base"]
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 110

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> None:
    ap = argparse.ArgumentParser(description="חוברת נתוני אמת סניף 500 (Power BI)")
    ap.add_argument(
        "--investigation-json",
        default="outputs/חקירת_סניף_500_2026-07-04.json",
    )
    ap.add_argument("--output-dir", default="outputs")
    args = ap.parse_args()

    inv_path = Path(args.investigation_json)
    investigation = None
    if inv_path.exists():
        investigation = json.loads(inv_path.read_text(encoding="utf-8"))

    out = Path(args.output_dir) / f"נתוני_אמת_סניף_500_{dt.date.today().isoformat()}.xlsx"
    path = build_workbook(investigation, out)
    print(f"[✓] נשמר: {path}")
    print(f"[i] אמת BI אגרגט סניף 500: {BI_TRUTH['aggregate_branch_500']['pct']:.2%}")
    print(f"[i] אמת BI יום 05/07: {BI_TRUTH['day_2026_07_05']['pct']:.2%}")
    print(f"[i] Power BI: {POWERBI_URL}")


if __name__ == "__main__":
    main()
